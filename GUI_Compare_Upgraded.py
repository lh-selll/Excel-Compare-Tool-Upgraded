# -*- coding: utf-8 -*-
## 基于DMEO3修改
## 增加.xlsm格式支持，添加Person_ComparisonApp类
## 增加当前task显示框及逻辑
## 文件对比功能基本修改完毕，
## 待办：保存文件步骤，用于处理结束的槽函数编写
## 待办：设置sheet或index表头列的选项list，选择文件后能够直接选择文件中的sheet和sheet中的title
## 待办：增加一列用于存放title所在的行
## 待办：不mapping title时，index能够分别选择两个文件的index_col，以解决两个文件中的index列不在同一列的问题 --（还是拆分两个表格会好一点，不会混乱，如果保持现状，可能需要在修改mapping title时重建表格，需调研一下）
## 
## 待办：增加当前task框，放入当前进度，停止按钮还是失效，需要修复
## 已完成增加当前task框，停止按钮问题已修复
## 待办：增加页面，用于展示增删点和修改点表格
## 取消展示增删点和修改点表格需求
## 2025/06/20：增加一键清空界面配置的功能
## 2025/06/21：增加一键清空界面配置的功能，增加一列，用于筛选新增、删除、修改的列
## 2025/06/21：完成增加一键清空界面配置的功能
## 2025/06/21：待办：增加一列，用于筛选新增、删除、修改的列
## 2025/06/22：已完成增加一列，用于筛选新增、删除、修改的列
## 2025/06/23：pyqt5->pyqt6>pyside6
## 2025/06/26：待办：增加初始化界面，包含一个开始按钮，点击开始按钮后进入主界面
## 2025/06/27：已完成增加初始化界面
## 2025/06/27：待办：解决改变表格行数时，sheet选项丢失的问题
## 2025/07/02：删除copy sheet，新增当前任务开始和结束的系统时间，解决progress回退问题
## 2025/07/03：优化界面布局，缩小高度
## 2025/07/03：取消license验证
## 2025/07/08 修复当索引列为空时，报“用户终止进程”的错误
## 2025/07/09 进一步压缩界面大小
## 2025/07/29：V16,对比结果框中添加"对比结果摘要"，同时将摘要放入对比结果log文件中，以快速浏览对比结果
## 2025/07/30：V16,增加三个按钮，分别可以快速打开对比结果excel文件，以及对比结果log文件
## 2025/07/30：V16,解决处理CSV文件时的问题
## 2025/07/31：待解决，直接对比时，底部空白行对比有差异
## 2025/08/02：已解决，打开CSV文件时格式报错问题
## 2025/08/02：待解决，第二个文件存在大量空行时，进度条存在问题，会从48%直接调至100%，继续测试json文件异常时的case
## 2025/08/02：已解决，第二个文件存在大量空行时，进度条存在的问题，
## 2025/08/02：待解决，继续测试json文件异常时的case
## 2025/08/03：已解决，读取csv文件时，会将空白格转换为“Nan”，导致对比差异的问题
## 2025/08/03：已解决，json文件异常时的case测试完成，增加在json中的mapping_status值异常时的处理策略，避免读取json文件时出现异常
##
## 

import ctypes
import sys
import os
import traceback
import inspect
import copy
import json
import time
# 记录应用程序启动时间（用于性能分析或日志记录）
system_start_time = time.time()
import openpyxl, xlrd
import textwrap
from collections import Counter
import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget,
    QTableWidget, QTableWidgetItem, QComboBox, QLineEdit, QPushButton, QLabel,
    QFileDialog, QMessageBox, QProgressBar, QHeaderView, QSpinBox, QCompleter,
    QPlainTextEdit, QSplashScreen, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QLayout, QAbstractScrollArea, QSizePolicy, QTextEdit
)
from PySide6.QtCore import Qt, QThread, Signal, QStringListModel, QSize
from PySide6.QtGui import QColor, QFont, QValidator, QPixmap, QPainter, QGuiApplication

from Person_ComparisonApp import Person_ComparisonApp
from Deviceid_license_verify import DeviceIDLicenseVerify
from FileHandler import FileHandler

output_path = ".\\outputfile"
json_file_path = '.\\json\\config.json'
license_file_path = '.\\license\\license.key'
compare_info_file_path = '.\\result.log'


class FileSelectorWidget(QWidget):
    """文件选择组件，包含标签、路径输入框和浏览按钮"""
    def __init__(self, label_text, path_edit_height, parent=None):
        super().__init__(parent)
        self.path_edit_height = path_edit_height
        self.name = label_text
        self.layout = QHBoxLayout(self)
        self.label = QLabel(label_text)
        # self.label.setMinimumHeight(30)
        self.path_edit = QLineEdit()
        self.path_edit.setReadOnly(True)
        self.path_edit.setMinimumWidth(180)
        self.path_edit.setMinimumHeight(self.path_edit_height+3)
        self.path_edit.setMaximumHeight(self.path_edit_height+3)
        self.browse_btn = QPushButton("浏览...")
        self.label.setStyleSheet("""
            QLabel {
                font-family: "SF Pro Display", "Helvetica Neue", Arial, sans-serif;
                font-size: 14px;
                font-weight: 800;
                color: #333333;
                border-color: #00E5EA;   
            }
        """)
        self.label.setAlignment(Qt.AlignCenter)
        self.path_edit.setStyleSheet("""
            QLineEdit {
                background-color: #F5F5F7;
                border: 1px solid #E5E5EA;
                border-radius: 10px;
                padding: 2px 4px;
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 10px;
                color: #333333;
                selection-background-color: #007AFF;
                selection-color: white;
            }
            QLineEdit:focus {
                border: 1px solid #007AFF;
                background-color: white;
                box-shadow: 0 0 0 3px rgba(0, 122, 255, 0.2);
            }
            QLineEdit:disabled {
                background-color: #F2F2F7;
                color: #A2A2A7;
                border-color: #E5E5EA;
            }
        """)
        self.browse_btn.setMinimumHeight(self.path_edit_height)
        self.browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #4B5563; /* 深灰背景，沉稳突出 */
                color: #FFFFFF; /* 白色文字，对比强烈 */
                border: 1px solid #6B7280; /* 深灰边框，统一色调 */
                border-radius: 8px;
                font-family: "Inter", "SF Pro Text", sans-serif;
                font-size: 13px;
                font-weight: 600;
                padding: 1px 1px;
                min-width: 80px;
                transition: all 0.2s ease;
            }
            QPushButton:hover {
                background-color: #6B7280; /* hover时深灰变浅，有呼吸感 */
                border-color: #9CA3AF;
                color: #FFFFFF;
            }
            QPushButton:pressed {
                background-color: #374151; /* 按下时深灰更暗，确认交互 */
                border-color: #4B5563;
                transform: translateY(1px);
            }
            QPushButton:disabled {
                background-color: #F5F5F5; 
                color: #BDBDBD;
                border-color: #E0E0E0;
            }
        """)
        
        self.layout.addWidget(self.label)
        self.layout.addWidget(self.path_edit, 1)
        self.layout.addWidget(self.browse_btn)
        self.layout.setAlignment(Qt.AlignCenter)
        
    def get_file_path(self):
        return self.path_edit.text()
    
    def set_file_path(self, path):
        self.path_edit.setText(path)

    def getName(self):
        """返回文件选择器的名称"""
        return self.name

class UpperCaseValidator(QValidator):
    """大写字母验证器"""
    def validate(self, input_text, pos):
        if input_text == "":
            return (QValidator.State.Acceptable, input_text, pos)
        if all(c.isalpha() and c.isupper() for c in input_text):
            return (QValidator.State.Acceptable, input_text, pos)
        return (QValidator.State.Invalid, input_text, pos)

class config_data_Container():
    """配置数据容器，存储单个对比任务的配置信息"""
    def __init__(self, sheet1_name="", sheet2_name="", mapping="N", title_row=1, col=["", "", ""]):
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name
        self.mapping = mapping
        self.col = col
        self.title_row = title_row
    def __repr__(self):
        return f"Container(s1={self.sheet1_name}, s2={self.sheet2_name}, mapping={self.mapping}, cols={self.col}, title_row={self.title_row})"

class restored_config_data_Container():
    """配置数据恢复容器，用于保存和加载配置数据到JSON文件"""
    def __init__(self, table_row_number):
        self.row_number = table_row_number  # 表格行数
        self.col_number = 7                 # 表格列数
        if isinstance(self.row_number, int) and self.row_number > 0:
            self.file1_path = ""
            self.file2_path = ""
            self.config_data = [config_data_Container() for _ in range(self.row_number)]

    def update_row_number(self, row_number):
        self.row_number = row_number
        self.config_data = None
        self.config_data = [config_data_Container() for _ in range(self.row_number)]

    def save_to_file(self, filename):
        """将配置数据保存到JSON文件"""
        if not os.path.exists(os.path.dirname(filename)):   #os.path.dirname(filename)获取文件的路径，并检查路径是否存在，
            os.makedirs(os.path.dirname(filename))  #路径不存在，创建路径
        with open(filename, 'w') as f:
            # 使用 json.dump 方法将字典数据写入文件
            json.dump(self.__dict__, f)

    def load_from_file(self, filename):
        """从JSON文件加载配置数据"""
        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            try:
                # 以读取模式打开文件
                with open(filename, 'r') as f:
                    # 使用 json.load 方法从文件中读取数据并转换为字典
                    data = json.load(f)
                    # 检查data长度是否符合要求
                    if data['row_number']  > 15:
                        print(f"jason文件中row_number={data['row_number']}，大于{15}，跳过读取jason文件操作。")
                        return 0
                    
                    self.row_number = data['row_number']

                    if len(data['config_data'])  != self.row_number:
                        print(f"jason文件中sheet_name_edit长度不等于{self.row_number}，跳过读取jason文件操作。")
                        return 0
                    else:
                        print(f"data = {data}")
                        for index in range(0, self.row_number):
                            if len(data['config_data'][index]) != self.col_number:
                                print(f"jason文件中config_data的第{index+1}行中的col的长度不等于{self.col_number-4}，跳过读取jason文件操作。")
                                return 0
                    # 从字典中获取相应的数据并更新存储类的属性
                    self.update_row_number(self.row_number)
                    self.__dict__.update(data)
                    print(f"update(data) = {self.__dict__}")
                    return 1
            except (FileNotFoundError, json.JSONDecodeError) as e:
                # 如果文件不存在，忽略错误，保持默认值
                print("文件不存在或JSONDecode错误，跳过读取操作。")
                return 0
        print("文件为空或不存在，跳过读取操作。")
        return 0
    
    def __repr__(self):
        return f"Container(row_number={self.row_number}, col_number={self.col_number}, file1_path={self.file1_path}, file2_path={self.file2_path}, config_data={self.config_data})"
        
class DataProcessor(QThread):
    """数据处理线程，负责Excel文件对比和数据处理"""
    progress_updated = Signal(int)          # 进度更新信号
    result_text_edit = Signal(str)              # 结果准备好信号
    progress_current_task = Signal(str)     # 当前任务进度信号
    comparison_finished = Signal(str)       # 对比完成信号
    error_occurred = Signal(str)            # 错误发生信号
    is_running = True                           # 线程运行状态

    def __init__(self, file1_path, file2_path, config_data, parent=None):
        super().__init__(parent)
        self.file1_path = file1_path            # 文件1路径
        self.file2_path = file2_path            # 文件2路径
        self.config_data = config_data          # 配置数据
        self.canceled = False                   # 取消标志
        self.index_col_position = [2, 4]         # 索引列位置范围
        self.CompareApp = Person_ComparisonApp(self.progress_updated, self.progress_current_task, self.comparison_finished)
        self.Thread_start_time = time.time()
        

    def run(self):
        """线程主函数，执行Excel对比流程"""
        current_progress_percent = 0    #当前进度条数值
        restored_config_data = restored_config_data_Container(len(self.config_data))
        restored_config_data.file1_path = self.file1_path
        restored_config_data.file2_path = self.file2_path
        restored_config_data.config_data = self.config_data
        restored_config_data.save_to_file(json_file_path)
        self.progress_current_task.emit("/*************************************************开始任务********************************************************/")
        timestamp = time.time()
        local_time = time.localtime(timestamp)
        formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", local_time)
        self.progress_current_task.emit(f"开始时间：{formatted_time}\nopenning File")

        try:
            self.progress_updated.emit(0)
            print(f"当前行数为：{inspect.currentframe().f_lineno}，DataProcessor")
            wb1, error_msg = self.open_file(self.file1_path)   # 打开文件1
            if not wb1:
                raise ValueError(f"打开文件1失败: {error_msg}")
            wb2, error_msg = self.open_file(self.file2_path)   # 打开文件2
            if not wb2:
                raise ValueError(f"打开文件2失败: {error_msg}")
            
            # 处理文件路径和输出路径
            file1_name, file1_ext = os.path.splitext(os.path.basename(self.file1_path))
            file2_name, file2_ext = os.path.splitext(os.path.basename(self.file2_path))
            if file1_ext.lower() == '.csv':
                # CSV文件默认转为XLSX格式
                file1_ext = '.xlsx'
            if file2_ext.lower() == '.csv':
                # CSV文件默认转为XLSX格式
                file2_ext = '.xlsx'
            output_path1 = os.path.join(output_path, f"{file1_name}-compare{file1_ext}")
            output_path2 = os.path.join(output_path, f"{file2_name}-compare{file2_ext}")

            # 处理配置数据
            results = []
            results_data = []
            self.progress_current_task.emit("开始获取配置表·····")
            for row in self.config_data:
                current_progress_percent += 2
                self.progress_updated.emit(current_progress_percent)
                if not row[0] or not row[1]:  # 跳过空行
                    continue
                
                # 解析配置行
                sheet1_name = row[0]
                sheet2_name = row[1]
                mapping = row[5] == 'Y'
                
                # 获取指定sheet
                wb1_sheet = wb1[sheet1_name]
                wb2_sheet = wb2[sheet2_name]
                
                if wb1_sheet is None:
                    raise ValueError(f"文件1中找不到Sheet: {sheet1_name}")
                if wb2_sheet is None:
                    raise ValueError(f"文件2中找不到Sheet: {sheet2_name}")
                
                data = config_data_Container(sheet1_name, sheet2_name, mapping, row[6], [])
                
                # 处理数据
                if mapping:  # Mapping title模式
                    for index in range(self.index_col_position[0], self.index_col_position[1]+1):
                        if not row[index]:
                            continue
                        else: 
                            data.col.append(row[index])
                    
                    # 这里添加实际的数据处理逻辑
                    result = f"处理映射: {sheet1_name} -> {sheet2_name}, 列: {data.col}"
                    results.append(result)
                else:  # No mapping模式
                    for index in range(self.index_col_position[0], self.index_col_position[1]+1):
                        if not row[index]:
                            continue
                        else:
                            # 根据索引列的列标号，转换为数字，例如：C转换后为3
                            index_value = self.get_index_by_ColHeader(row[index])
                            if index_value > wb1_sheet.max_column and index_value > wb2_sheet.max_column:
                                raise ValueError(f"第{row}行的第{index+1}列输入值超出最大值{min(wb1_sheet.max_column, wb2_sheet.max_column)}: {sheet1_name}")
                            else:
                                data.col.append(index_value)
                    
                    # 这里添加实际的数据处理逻辑
                    result = f"处理索引: {sheet1_name}  ->  {sheet2_name}, 索引: {data.col}"
                    results.append(result)
                results_data.append(data)
            self.progress_current_task.emit("完成获取配置表，开始对比")

            # 计算每完成一列的进度步进
            self.progress_current_task.emit(f"results_data数据整理完成：【 {results_data}】")
            delta_progress = float((90 - current_progress_percent)/len(results_data)/2)
            
            compare_result_info = ""
            for row_data in results_data:
                wb1_sheet = wb1[row_data.sheet1_name]
                status, error_msg = self.CompareApp.delete_bottom_blank_rows(wb1_sheet)
                if not status:
                    raise ValueError(f"删除{row_data.sheet1_name}底行失败: {error_msg}")
                wb2_sheet = wb2[row_data.sheet2_name]
                status, error_msg = self.CompareApp.delete_bottom_blank_rows(wb2_sheet)
                if not status:
                    raise ValueError(f"删除{row_data.sheet2_name}底行失败: {error_msg}")
                # 更新当前进度
                wb1_sheet_copy = wb1.copy_worksheet(wb1_sheet)
                if not row_data.mapping and len(row_data.col) == 0:
                    # 直接对比 mapping=0, 未填写index列数
                    print(f"当前行数为：{inspect.currentframe().f_lineno} compare_excel_sheet")
                    if not self.CompareApp.compare_excel_sheet(wb1_sheet, wb2_sheet, current_progress_percent, current_progress_percent+delta_progress):
                        raise ValueError(f"用户终止对比进程")
                    current_progress_percent += delta_progress
                    compare_result_info += f"Sheet Name: {row_data.sheet1_name} -> {row_data.sheet2_name}\n===============文件1相比文件2==============={self.CompareApp.result_info}"
                    
                    if not self.CompareApp.compare_excel_sheet(wb2_sheet, wb1_sheet_copy, current_progress_percent, current_progress_percent+delta_progress):
                        raise ValueError(f"用户终止对比进程")
                    current_progress_percent += delta_progress
                    compare_result_info += f"===============文件2相比文件1==============={self.CompareApp.result_info}\n"
                elif not row_data.mapping and len(row_data.col) != 0:
                    # 根据索引值对比， mapping=0, 填写index列数
                    print(f"当前行数为：{inspect.currentframe().f_lineno} compare_excel_sheet_by_index, row_data.col = {row_data.col}")
                    status, add_sheet1 = self.CompareApp.compare_excel_sheet_by_index(wb1_sheet, wb2_sheet, row_data.col, file1_name, current_progress_percent, current_progress_percent+delta_progress)
                    if not status:
                        raise ValueError(f"用户终止对比进程")
                    current_progress_percent += delta_progress
                    compare_result_info += f"Sheet Name: {row_data.sheet1_name} -> {row_data.sheet2_name}\n===============文件1相比文件2==============={self.CompareApp.result_info}"
                    status, add_sheet2 = self.CompareApp.compare_excel_sheet_by_index(wb2_sheet, wb1_sheet_copy, row_data.col, file2_name, current_progress_percent, current_progress_percent+delta_progress)
                    if not status:
                        raise ValueError(f"用户终止对比进程")
                    self.CompareApp.merge_sheet_to_another(add_sheet2, wb1_sheet)
                    self.CompareApp.merge_sheet_to_another(add_sheet1, wb2_sheet)
                    current_progress_percent += delta_progress
                    compare_result_info += f"===============文件2相比文件1==============={self.CompareApp.result_info}\n"
                else:
                    # 按索引和表头映射对比, mapping=1, 填写index列数
                    print(f"当前行数为：{inspect.currentframe().f_lineno} compare_excel_sheet_by_index_mapping_title, row_data.col = {row_data.col}")
                    status, add_sheet1 = self.CompareApp.compare_excel_sheet_by_index_mapping_title(wb1_sheet, wb2_sheet, row_data.col, data.title_row, file1_name, current_progress_percent, current_progress_percent+delta_progress)
                    if not status:
                        raise ValueError(f"用户终止对比进程")
                    print(f"当前行数为：{inspect.currentframe().f_lineno} compare_excel_sheet")
                    
                    self.progress_current_task.emit(f"当前行数为：{inspect.currentframe().f_lineno} 对比第二个文件-------------")
                    current_progress_percent += delta_progress
                    compare_result_info += f"Sheet Name: {row_data.sheet1_name} -> {row_data.sheet2_name}\n===============文件1相比文件2==============={self.CompareApp.result_info}"
                    status, add_sheet2 = self.CompareApp.compare_excel_sheet_by_index_mapping_title(wb2_sheet, wb1_sheet_copy, row_data.col, data.title_row, file2_name, current_progress_percent, current_progress_percent+delta_progress)
                    if not status:
                        raise ValueError(f"用户终止对比进程")
                    self.CompareApp.merge_sheet_to_another(add_sheet2, wb1_sheet)
                    self.CompareApp.merge_sheet_to_another(add_sheet1, wb2_sheet)
                    current_progress_percent += delta_progress
                    compare_result_info += f"===============文件2相比文件1==============={self.CompareApp.result_info}\n"
                # 删除工作表
                wb1.remove(wb1_sheet_copy)
            
            # 保存对比结果
            compare_compeleted_time = time.time()
            self.progress_current_task.emit("完成所有sheet对比任务，开始保存File")
            self.progress_updated.emit(90)
            self.progress_current_task.emit("完成所有sheet对比任务，开始保存File")
            if self.saving_file(wb1, output_path1):
                self.progress_current_task.emit("File1保存成功")
                self.progress_current_task.emit(f"File1:output_path1 = {output_path1}")
            self.progress_updated.emit(95)
            if self.saving_file(wb2, output_path2):
                self.progress_current_task.emit("File2保存成功")
                self.progress_current_task.emit(f"File2:output_path2 = {output_path2}")
            self.progress_current_task.emit("成功保存所有文档")
            wb1.close()
            wb2.close()
            saving_compeleted_time = time.time()
            self.progress_current_task.emit(textwrap.dedent(f"""
            ======================================
            完成所有sheet对比任务耗时：{compare_compeleted_time - self.Thread_start_time}s
            保存所有文档耗时：{saving_compeleted_time - compare_compeleted_time}s
            本次任务总耗时：{time.time()-self.Thread_start_time}s
            ======================================
            """))

            # 生成结果报告
            report = "数据处理完成!\n"
            report += f"文件1: {os.path.basename(self.file1_path)}\n"
            report += f"文件2: {os.path.basename(self.file2_path)}\n"
            report += "处理结果:\n"
            report += "\n".join(results)
            report += "\n\n"+ compare_result_info
            
            self.result_text_edit.emit(report)
            FileHandler.delete_file(compare_info_file_path)
            FileHandler.create_text_file(compare_info_file_path)
            if not FileHandler.append_text_content(compare_info_file_path, report):
                self.progress_current_task.emit(f"{compare_info_file_path}文件追加错误")

            self.comparison_finished.emit("success")
            self.progress_updated.emit(100)
            timestamp = time.time()
            local_time = time.localtime(timestamp)
            formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", local_time)
            self.progress_current_task.emit(f"结束时间：{formatted_time}")
            self.progress_current_task.emit("/*************************************************结束任务*******************************************************/")
            ctypes.windll.user32.MessageBoxW(None, f"对比完成，输出文件在“{output_path}”文件夹中", "成功信息", 0x00000040)
        except Exception as e:
            self.error_occurred.emit(f"处理失败: {str(e)}")
            self.comparison_finished.emit("failed")

            # self.progress_updated.emit(0)
        # finally:
        #     self.progress_updated.emit(100)
        
        return None

    def stop(self):
        """停止对比进程"""
        self.CompareApp.is_running = False

    def get_index_by_ColHeader(self, ColumnHeader):
        """将列字母转换为列号（如C->3）"""
        try:
            # 存放index的计算结果
            index_value = 0
            for s in range(0, len(ColumnHeader)):
                index_arrays = ord(ColumnHeader[s]) - ord('A') + 1 #计算每个字母的值
                index_value = index_value + index_arrays * (26 ** (len(ColumnHeader) - s - 1))
            print(f"index_value = {index_value}")
            return index_value
        except ValueError:
            error = f"get_index_by_ColHeader 失败"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
    @staticmethod
    def open_file(file_path, read_only_flag = False):
        """打开Excel文件，支持.xls/.xlsx/.xlsm/.csv格式"""
        # 加载一个 Excel 文件
        try:
            if file_path.lower().endswith('.xls'):
                # 处理 .xls 文件
                wb = openpyxl.Workbook()
                xls_wb = xlrd.open_workbook(file_path, on_demand=read_only_flag)
                for sheet_name in xls_wb.sheet_names():
                    xls_sheet = xls_wb.sheet_by_name(sheet_name)
                    new_sheet = wb.create_sheet(sheet_name)
                    for row in range(xls_sheet.nrows):
                        for col in range(xls_sheet.ncols):
                            new_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)
                del wb['Sheet']  # 删除默认创建的工作表
            elif file_path.lower().endswith('.csv'):
                try:
                    # 尝试常见编码格式
                    encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'gb18030']
                    for encoding in encodings:
                        try:
                            df = pd.read_csv(
                                file_path,
                                encoding=encoding,
                                keep_default_na=False,  # 关键：不将空白格解析为NaN
                                na_values=[]  # 额外确保不将任何值识别为NaN（可选）
                            )
                            df = df.fillna("")
                            print(f"成功读取文件，使用编码: {encoding}")
                            break
                        except UnicodeDecodeError:
                            if encoding == encodings[-1]:
                                return None, "无法识别文件编码，请检查文件格式"
                            else:
                                continue
                
                except Exception as e:
                    return None, f"读取文件出错: {str(e)}"
                # 直接创建openpyxl工作簿并写入数据（无临时文件）
                wb = openpyxl.Workbook()
                ws = wb.active
                # 写入表头
                ws.append(df.columns.tolist())
                # 写入数据行
                for row in df.itertuples(index=False, name=None):
                    ws.append(row)
            elif file_path.lower().endswith('.xlsx'):
                # 处理 .xlsx 文件
                wb = openpyxl.load_workbook(file_path, read_only=read_only_flag, data_only=True)
            
            elif file_path.lower().endswith('.xlsm'):
                # 处理 .xlsm 文件
                wb = openpyxl.load_workbook(file_path, keep_vba=True, read_only=read_only_flag, data_only=True)
                
        except FileNotFoundError:
            error = f"文件 {file_path} 不存在。"
            print(error)
            # ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return (None, error)
        except openpyxl.utils.exceptions.InvalidFileException:
            error = f"文件 {file_path} 不是有效的 Excel 文件, 请重新输入"
            print(error)
            # ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return (None, error)
        except Exception as e:
            error = f"发生了未知错误：{e}"
            print(error)
            # ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return (None, error)
        return (wb, None)

    def saving_file(self, wb, output_path):
        """保存Excel文件"""
        try:
            # 根据文件扩展名选择保存方式
            ext = os.path.splitext(output_path)[1].lower()
            
            if ext in ('.xlsx', '.xlsm', '.xls'):
                # 使用 openpyxl 保存
                wb.save(output_path)
            
            elif ext == '.csv':
                # 保存为 CSV（适合用 pandas 处理）
                import pandas as pd
                df = pd.DataFrame(wb.active.values)  # 从工作表提取数据
                df.to_csv(output_path, index=False, header=False)
            
            print(f"文件保存成功: {output_path}")
            return 1
        except Exception as e:
            if isinstance(e, PermissionError):
                error = f"没有权限保存文件到指定路径，请检查文件权限设置。"
            elif isinstance(e, OSError) and "磁盘空间不足" in str(e):
                error = f"磁盘空间不足，无法保存文件，请清理磁盘空间后再试。"
            elif isinstance(e, FileNotFoundError):
                error = f"保存文件时文件路径不存在：{str(e)}"
                try:
                    output_dir = os.path.dirname(output_path)
                    os.mkdir(output_dir)
                    error = f"文件夹 {output_dir} 创建成功。"
                    # 根据文件扩展名选择保存方式
                    ext = os.path.splitext(output_path)[1].lower()
                    
                    if ext in ('.xlsx', '.xlsm', '.xls'):
                        # 使用 openpyxl 保存
                        wb.save(output_path)
                    
                    elif ext == '.csv':
                        # 保存为 CSV（适合用 pandas 处理）
                        import pandas as pd
                        df = pd.DataFrame(wb.active.values)  # 从工作表提取数据
                        df.to_csv(output_path, index=False, header=False)
                    
                    print(f"文件保存成功: {output_path}")
                    return 1
                except FileExistsError:
                    error = f"文件夹 {output_path} 已经存在。"
                except PermissionError:
                    error = f"没有权限创建文件夹 {output_path}。"
            else:
                error = f"保存文件时出现未知错误：{str(e)}"
            print(error)
            # ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            self.progress_current_task.emit(f"{error}")
            raise ValueError(error)
    
class DataProcessingTool(QMainWindow):
    """主应用程序窗口"""
    global output_path
    global compare_info_file_path
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel对比工具")
        self.table_row_number = 3               #输入table的总行数
        self.table_column_number = 7            #输入table的总列数
        self.index_col_position = [2, 4]        #放置index列位置，比如第2列到第4列，方便后续扩展（从0开始）
        self.table_row_number_range = [3, 16]    #放置index列位置，比如第2列到第4列，方便后续扩展（从0开始）
        self.mapping_option = 5                 #选择是否需要mapping title选项所在的列，（从0开始）
        self.title_rows = 6                     #title列所在的列数（从0开始）
        self.table_row_height = 32              # 表格行高
        self.table_heigh = self.table_row_number * self.table_row_height+50  # table的高度
        self.current_task_label_height = 130     #当前任务展示框的高度
        self.select_edit_heigh = 15             #index列选项框的高度
        self.progress_bar_heigh = 35            #进度条和开始按钮的高度
        self.path_edit_height = 23              #文件选项框的高度
        # self.setGeometry(400, 100, 1000, 600)
        setup_window_geometry(self, 1000, 20)
        self.output_file_path1 = ""
        self.output_file_path2 = ""
        
        # 当前配置数据
        self.config_data = [[] for _ in range(self.table_row_number)]
        self.wb1 = None
        self.wb2 = None
        self.title_list = []
        self.start_flag = True
        self.Button_Color = "#CECECE"      # 按钮默认颜色
        self.Button_click_Color = "#079E61" # 按钮点击颜色
        self.Border_color = "#4480b2"      # 边框颜色
        
        # 初始化UI
        self.init_ui()

    def init_ui(self):
        global system_start_time

        """初始化用户界面"""
        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)
        self.file_operator = FileHandler(self)
        
        # 一键清空配置按钮
        clear_button_layout = QHBoxLayout()
        self.clear_button = QPushButton("One-click clear")
        self.clear_button.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                        stop:0 #616161,
                                        stop:1 #424242);
                color: white;
                border-radius: 12px;
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 16px;
                font-weight: 800;
                padding: 5px 12px;
                border: none;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.2),
                            0 1px 2px rgba(0, 0, 0, 0.1),
                            inset 0 1px 0 rgba(255, 255, 255, 0.1);
                transition: all 0.2s ease;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                        stop:0 #757575,
                                        stop:1 #545454);
                box-shadow: 0 4px 10px rgba(0, 0, 0, 0.25),
                            0 2px 4px rgba(0, 0, 0, 0.15),
                            inset 0 1px 0 rgba(255, 255, 255, 0.15);
                transform: translateY(-1px);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                        stop:0 #424242,
                                        stop:1 #212121);
                box-shadow: 0 1px 3px rgba(0, 0, 0, 0.2),
                            0 0 2px rgba(0, 0, 0, 0.1),
                            inset 0 1px 2px rgba(0, 0, 0, 0.1);
                transform: translateY(1px);
            }
            QPushButton:disabled {
                background: #E0E0E0;
                color: #9E9E9E;
                box-shadow: none;
                border: 1px solid #BDBDBD;
            }
        """)

        self.clear_button.clicked.connect(self.One_click_clear)
        self.clear_button.setFixedHeight(30)
        self.clear_button.setMinimumWidth(800)
        clear_button_layout.addWidget(self.clear_button)

        # 文件选择区域
        file_layout = QVBoxLayout()
        self.file1_selector = FileSelectorWidget("文件1:", self.path_edit_height)
        self.file1_selector.browse_btn.clicked.connect(lambda: self.browse_file(self.file1_selector))
        self.file2_selector = FileSelectorWidget("文件2:", self.path_edit_height)
        self.file2_selector.browse_btn.clicked.connect(lambda: self.browse_file(self.file2_selector))
        file_layout.addWidget(self.file1_selector)
        file_layout.addWidget(self.file2_selector)
        file_layout.setAlignment(Qt.AlignTop)
        

        # 选择需要对比的sheet数量
        self.table_row_number_layout = QHBoxLayout()
        table_row_number_label = QLabel("选择需要对比的sheet数量")
        table_row_number_label.setFixedHeight(self.select_edit_heigh+15)
        table_row_number_label.setStyleSheet("""
            QLabel {
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 13px;
                font-weight: 700;
                color: #333333;
            }
        """)
        table_row_number_label.setAlignment(Qt.AlignCenter| Qt.AlignRight)
        table_row_number_label.setFixedWidth(300)  # 增加下拉框高度
        self.table_row_number_combo = QComboBox()
        self.table_row_number_combo.setMinimumHeight(self.select_edit_heigh+15)  # 增加下拉框高度
        self.table_row_number_combo.setFixedWidth(120)
        self.table_row_number_combo.addItems([str(i) for i in range(self.table_row_number_range[0], self.table_row_number_range[1])])
        index_list = self.get_combo_all_options(self.table_row_number_combo)
        index = index_list.index(str(self.table_row_number))
        self.table_row_number_combo.setCurrentIndex(index)
        self.table_row_number_combo.currentIndexChanged.connect(self.table_row_number_changed)
        self.table_row_number_layout.addWidget(table_row_number_label, alignment=Qt.AlignRight)
        self.table_row_number_layout.addWidget(self.table_row_number_combo, alignment=Qt.AlignRight)
        self.table_row_number_layout.setAlignment(Qt.AlignTop| Qt.AlignRight)
        # 设置 QComboBox 可编辑
        self.table_row_number_combo.setEditable(True)
        # 获取 QComboBox 内部的 QLineEdit 对象
        line_edit = self.table_row_number_combo.lineEdit()
        # 设置文字居中对齐
        line_edit.setAlignment(Qt.AlignCenter)

        # 选项卡区域
        self.tab_widget = QTabWidget()
        # 对比配置表输入框
        self.Compare_Config = QWidget()
        self.Compare_Config.setStyleSheet("""
            font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
            font-size: 10px;       /* 设置字体大小为14px */
            font-weight: 400;      /* 设置字体粗细为中等（500） */
        """)
        self.init_no_mapping_tab()
        self.tab_widget.addTab(self.Compare_Config, "Compare Config")
        # 创建一个水平布局来放置选项卡和下拉框
        tab_combo_layout = QHBoxLayout()
        tab_combo_layout.addWidget(self.table_row_number_combo)
        tab_combo_layout.setAlignment(Qt.AlignTop)
        
        # 进度和按钮区域
        progress_bar_layout = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        # 设置进度条样式
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                /* 进度条背景 */
                background-color: #F5F5F7;
                border-radius: 8px;
                height: 12px;
                text-align: center;
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 12px;
                color: #666666;
            }
            QProgressBar::chunk {
                /* 进度条填充部分 - 苹果蓝渐变 */
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                           stop:0 #007AFF,
                                           stop:1 #34C759);  /* 从蓝到绿的渐变，符合苹果风格 */
                border-radius: 8px;
            }
        """)
        self.progress_bar.setMinimumHeight(self.progress_bar_heigh)
        self.progress_bar.setMaximumHeight(self.progress_bar_heigh)
        
        # 开始按钮
        self.start_btn = QPushButton("开始处理")
        self.start_btn.clicked.connect(self.start_comparison)
        self.start_btn.setEnabled(False)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(148, 217, 255, 0.95),
                                           stop:1 #007AFF);
                color: white;
                border-radius: 12px;
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 16px;
                font-weight: 800;
                padding: 3px 6px;
                border: none;
                box-shadow: 0 3px 8px rgba(0, 122, 255, 0.2);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(180, 228, 255, 0.98),
                                           stop:1 #0066CC);
                box-shadow: 0 4px 10px rgba(0, 122, 255, 0.3);
                transform: translateY(-1px);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(100, 180, 255, 0.9),
                                           stop:1 #005299);
                box-shadow: 0 1px 4px rgba(0, 122, 255, 0.1);
                transform: translateY(1px);
            }
            QPushButton:disabled {
                background: #E5E5EA;
                color: #A2A2A7;
                box-shadow: none;
            }
        """)
        self.start_btn.setMinimumHeight(self.progress_bar_heigh)
        self.start_btn.setMaximumHeight(self.progress_bar_heigh)
        progress_bar_layout.addWidget(self.progress_bar, 4)
        progress_bar_layout.addWidget(self.start_btn, 1)
        progress_bar_layout.setAlignment(Qt.AlignTop)

        # 当前Task框
        current_task_layout = QHBoxLayout()
        current_task_label_layout = QHBoxLayout()
        current_task_label = QLabel("Print Info")
        current_task_label.setFixedHeight(20)
        current_task_label.setStyleSheet("""
            QLabel {
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 12px;
                font-weight: 800;
                color: #333333;
            }
        """)
        self.current_task_edit = QPlainTextEdit ()
        self.current_task_edit.setMinimumHeight(self.current_task_label_height)
        self.current_task_edit.setStyleSheet("""
            QPlainTextEdit {
                background-color: #FFFFFF;
                border: 1px solid #E5E5EA;
                border-radius: 12px;
                padding: 12px;
                font-family: "SF Mono", "Menlo", "Consolas", "Monaco", monospace;
                font-size: 12px;
                color: #333333;
                selection-background-color: rgba(0, 122, 255, 0.15);
                selection-color: #007AFF;
                outline: none;
            }
            
            /* 滚动条样式 */
            QScrollBar:vertical {
                border: none;
                background: #F5F5F7;
                width: 10px;
                margin: 0;
                border-radius: 5px;
            }
            
            QScrollBar::handle:vertical {
                background: #C7C7CC;
                min-height: 30px;
                border-radius: 5px;
            }
            
            QScrollBar::handle:vertical:hover {
                background: #A2A2A7;
            }
            
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0;
                subcontrol-position: bottom;
                subcontrol-origin: margin;
            }
            
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
        """)
        self.current_task_edit.setReadOnly(True)
        current_task_label_layout.addWidget(current_task_label)
        current_task_label_layout.setAlignment(Qt.AlignTop)
        current_task_layout.addWidget(self.current_task_edit)
        current_task_layout.setAlignment(Qt.AlignTop)
        # 创建按钮的垂直布局
        button_layout = QVBoxLayout()
        # 创建上方按钮
        self.button_up = QPushButton("文件1")
        button_layout.addWidget(self.button_up)
        self.button_up.setStyleSheet("""
            QPushButton {
                background-color: #e1e3e1; /* 深灰背景，沉稳突出 */
                color: #000000; /* 黑色文字，对比强烈 */
                border: 1px solid #e1e3e1; /* 深灰边框，统一色调 */
                border-radius: 12px;
                font-family: "Inter", "SF Pro Text", sans-serif;
                font-size: 15px;
                font-weight: 600;
                padding: 1px 1px;
                min-width: 80px;
                transition: all 0.2s ease;
            }
            QPushButton:hover {
                background-color: #7a7a7c; /* hover时深灰变浅，有呼吸感 */
                border-color: #9CA3AF;
                color: #FFFFFF;
            }
            QPushButton:pressed {
                background-color: #374151; /* 按下时深灰更暗，确认交互 */
                border-color: #4B5563;
                transform: translateY(1px);
            }
            QPushButton:disabled {
                background-color: #F5F5F5; 
                color: #BDBDBD;
                border-color: #E0E0E0;
            }
        """)
        self.button_up.setMinimumHeight(self.current_task_label_height/2-4)
        self.button_up.clicked.connect(lambda: self.open_result_file(0))
        self.button_up.setEnabled(0)
        # 创建下方按钮
        self.button_down = QPushButton("文件2")
        button_layout.addWidget(self.button_down)
        self.button_down.setStyleSheet("""
            QPushButton {
                background-color: #e1e3e1; /* 深灰背景，沉稳突出 */
                color: #000000; /* 黑色文字，对比强烈 */
                border: 1px solid #e1e3e1; /* 深灰边框，统一色调 */
                border-radius: 12px;
                font-family: "Inter", "SF Pro Text", sans-serif;
                font-size: 15px;
                font-weight: 600;
                padding: 1px 1px;
                min-width: 80px;
                transition: all 0.2s ease;
            }
            QPushButton:hover {
                background-color: #7a7a7c; /* hover时深灰变浅，有呼吸感 */
                border-color: #9CA3AF;
                color: #FFFFFF;
            }
            QPushButton:pressed {
                background-color: #374151; /* 按下时深灰更暗，确认交互 */
                border-color: #4B5563;
                transform: translateY(1px);
            }
            QPushButton:disabled {
                background-color: #F5F5F5; 
                color: #BDBDBD;
                border-color: #E0E0E0;
            }
        """)
        self.button_down.setMinimumHeight(self.current_task_label_height/2-4)
        self.button_down.clicked.connect(lambda: self.open_result_file(1))
        self.button_down.setEnabled(0)
        # 设置按钮在水平和垂直方向都可拉伸
        self.button_up.setSizePolicy(
            QSizePolicy.Preferred,  # 水平方向：尽可能填充可用空间
            QSizePolicy.Expanding   # 垂直方向：尽可能填充可用空间
        )
        self.button_down.setSizePolicy(
            QSizePolicy.Preferred,  # 水平方向：尽可能填充可用空间
            QSizePolicy.Expanding   # 垂直方向：尽可能填充可用空间
        )
        button_layout.setAlignment(Qt.AlignTop)
        current_task_layout.addLayout(button_layout)

        # 结果区域框
        result_layout = QHBoxLayout()
        self.result_text_edit = QTextEdit()
        self.result_text_edit.setStyleSheet("""
            QPlainTextEdit {
                background-color: #FFFFFF;
                border: 1px solid #E5E5EA;
                border-radius: 12px;
                padding: 14px;
                font-family: "SF Mono", "Menlo", "Consolas", "Monaco", monospace;
                font-size: 12px;
                font-weight: 1200;
                color: #333333;
                selection-background-color: rgba(0, 122, 255, 0.15);
                selection-color: #007AFF;
                outline: none;
            }
            
            /* 滚动条样式 */
            QScrollBar:vertical {
                border: none;
                background: #F5F5F7;
                width: 10px;
                margin: 0;
                border-radius: 5px;
            }
            
            QScrollBar::handle:vertical {
                background: #C7C7CC;
                min-height: 30px;
                border-radius: 5px;
            }
            
            QScrollBar::handle:vertical:hover {
                background: #A2A2A7;
            }
            
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0;
                subcontrol-position: bottom;
                subcontrol-origin: margin;
            }
            
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
        """)
        self.result_text_edit.setMinimumHeight(self.current_task_label_height-40)
        self.result_text_edit.setReadOnly(True)
        self.button_log = QPushButton("result.log")
        self.button_log.setStyleSheet("""
            QPushButton {
                background-color: #e1e3e1; /* 深灰背景，沉稳突出 */
                color: #000000; /* 黑色文字，对比强烈 */
                border: 1px solid #e1e3e1; /* 深灰边框，统一色调 */
                border-radius: 12px;
                font-family: "Inter", "SF Pro Text", sans-serif;
                font-size: 15px;
                font-weight: 600;
                padding: 1px 1px;
                min-width: 80px;
                transition: all 0.2s ease;
            }
            QPushButton:hover {
                background-color: #7a7a7c; /* hover时深灰变浅，有呼吸感 */
                border-color: #9CA3AF;
                color: #FFFFFF;
            }
            QPushButton:pressed {
                background-color: #374151; /* 按下时深灰更暗，确认交互 */
                border-color: #4B5563;
                transform: translateY(1px);
            }
            QPushButton:disabled {
                background-color: #F5F5F5; 
                color: #BDBDBD;
                border-color: #E0E0E0;
            }
        """)
        self.button_log.setMinimumHeight(self.current_task_label_height-50)
        self.button_log.clicked.connect(self.open_log_file)
        self.button_log.setEnabled(0)
        # 设置按钮在水平和垂直方向都可拉伸
        self.button_log.setSizePolicy(
            QSizePolicy.Preferred,  # 水平方向：尽可能填充可用空间
            QSizePolicy.Expanding   # 垂直方向：尽可能填充可用空间
        )
        result_layout.addWidget(self.result_text_edit)
        result_layout.addWidget(self.button_log)
        
        # 添加到主布局
        main_layout.addLayout(clear_button_layout)
        main_layout.addLayout(file_layout)
        main_layout.addLayout(self.table_row_number_layout)
        main_layout.addWidget(self.tab_widget, alignment=Qt.AlignTop)
        main_layout.addLayout(progress_bar_layout)
        main_layout.addLayout(current_task_label_layout)
        main_layout.addLayout(current_task_layout)
        main_layout.addLayout(result_layout)
        main_layout.setAlignment(progress_bar_layout, Qt.AlignTop)
        main_layout.setAlignment(Qt.AlignTop)
        # 在底部添加弹簧，但权重较小，允许窗口缩小
        main_layout.addStretch(0.01)
        
        # 连接文件选择信号
        self.file1_selector.path_edit.textChanged.connect(self.check_files_selected)
        self.file2_selector.path_edit.textChanged.connect(self.check_files_selected)
        self.setCentralWidget(main_widget)
        self.restored_config_data = restored_config_data_Container(15)
        if not self.restored_config_data.load_from_file(json_file_path):
            self.restored_config_data.update_row_number(self.table_row_number)
            if os.path.exists(json_file_path):
                os.remove(json_file_path)  # 删除文件
                print(f"文件 {json_file_path} 已成功删除")
        else:
            restored_result, error_msg = self.restore_current_data(self.restored_config_data) #加载历史数据
            if not restored_result:
                self.current_task_edit.appendPlainText(f"配置文件有误，已清空配置：\n原因:{error_msg}")
                if os.path.exists(json_file_path):
                    os.remove(json_file_path)  # 删除文件
                    print(f"文件 {json_file_path} 已成功删除")
            else:
                self.current_task_edit.appendPlainText(f"初始化完成，恢复历史配置")

        delta_time = time.time() - system_start_time
        print(f"启动时间 = {delta_time}s")
        self.current_task_edit.appendPlainText(f"启动时间 = {delta_time}s")

    def restore_current_data(self, restored_data):
        """恢复历史配置数据"""
        try:
            print(f"func: restore_current_data, restored_data = {restored_data}")
            file1_path = restored_data.file1_path
            file2_path = restored_data.file2_path
            self.table_row_number = restored_data.row_number
            
            # 获取下拉框选项并设置索引
            index_list = self.get_combo_all_options(self.table_row_number_combo)
            try:
                index = index_list.index(str(self.table_row_number))
            except ValueError:
                return (0, f"找不到与行数 {self.table_row_number} 匹配的下拉选项")
            self.table_row_number_combo.setCurrentIndex(index)
            
            # 检查文件路径是否存在
            if not (file1_path and file2_path):
                error_msg = f"文件路径不完整: file1={file1_path}, file2={file2_path}"
                print(error_msg)
                return (0, error_msg)
            
            # 打开第一个文件
            wb1, error_msg = DataProcessor.open_file(file1_path, True)
            if not wb1:
                return (0, f"打开文件1失败: {error_msg}")
            
            # 打开第二个文件
            wb2, error_msg = DataProcessor.open_file(file2_path, True)
            if not wb2:
                return (0, f"打开文件2失败: {error_msg}")
            
            # 设置文件路径和工作簿
            self.file1_selector.set_file_path(file1_path)
            self.file2_selector.set_file_path(file2_path)
            self.output_file_path1 = self.get_file_output_path_byFilepath(file1_path)
            self.output_file_path2 = self.get_file_output_path_byFilepath(file2_path)
            self.wb1 = wb1
            self.wb2 = wb2
            print(f"wb1.sheetnames = {wb1.sheetnames}")
            print(f"wb2.sheetnames = {wb2.sheetnames}")
            
            # 处理表格配置数据
            for row in range(0, self.table_row_number):
                try:
                    # 提取配置数据
                    sheet1_name = restored_data.config_data[row][0]
                    sheet2_name = restored_data.config_data[row][1]
                    mapping_status = restored_data.config_data[row][self.mapping_option]
                    index_column_list = restored_data.config_data[row][self.index_col_position[0]: self.index_col_position[1]+1]
                    title_row_number = restored_data.config_data[row][self.title_rows]
                    
                    print(f"sheet1_name = {sheet1_name}, sheet2_name = {sheet2_name}")
                    
                    # 获取工作表索引
                    try:
                        index1 = list(self.wb1.sheetnames).index(sheet1_name) + 1 if sheet1_name else 0
                        index2 = list(self.wb2.sheetnames).index(sheet2_name) + 1 if sheet2_name else 0
                    except ValueError:
                        return (0, f"工作表不存在: sheet1={sheet1_name}, sheet2={sheet2_name}")
                    
                    # 设置第一个工作表下拉框
                    sheet_combo = self.Compare_Config_table.cellWidget(row, 0)
                    sheet_combo.clear()
                    sheet_combo.addItems([""] + self.wb1.sheetnames)
                    sheet_combo.setCurrentIndex(index1)
                    
                    # 设置第二个工作表下拉框
                    sheet_combo = self.Compare_Config_table.cellWidget(row, 1)
                    sheet_combo.clear()
                    sheet_combo.addItems([""] + self.wb2.sheetnames)
                    sheet_combo.setCurrentIndex(index2)
                            
                    # 设置映射状态
                    mapping_status_combo = self.Compare_Config_table.cellWidget(row, self.mapping_option)
                    if mapping_status not in ["Y", "N"]:
                        mapping_status_combo.setCurrentText("N")
                        print(f"未知映射状态mapping_status: {mapping_status}")
                        mapping_status_combo.setCurrentText("N")  # 默认设置为N
                        continue  # 跳过当前循环
                    mapping_status_combo.setCurrentText(mapping_status)
                    
                    # 处理映射配置
                    if sheet1_name and sheet2_name:
                        sheet1 = self.wb1[sheet1_name]
                        sheet2 = self.wb2[sheet2_name]
                        title_list = self.get_title_list(sheet1, sheet2, title_row_number)
                        
                        # 处理索引列配置
                        for col in range(self.index_col_position[0], self.index_col_position[1]+1):
                            try:
                                index = title_list.index(index_column_list[col-2]) + 1 if index_column_list[col-2] else 0
                            except ValueError:
                                index = 0  # 未找到匹配项时设置为0
                            
                            # 根据映射状态设置其他控件
                            if mapping_status_combo.currentText() == "Y":
                                mapping_row_combo = self.Compare_Config_table.cellWidget(row, self.title_rows)
                                mapping_row_combo.setValue(title_row_number)
                                index_combo = self.Compare_Config_table.cellWidget(row, col)
                                index_combo.setCurrentIndex(index)
                            elif mapping_status_combo.currentText() == "N":
                                index_combo = self.Compare_Config_table.cellWidget(row, col)
                                index_combo.clear()
                                index_combo.setText(index_column_list[col-2] or "")
                            else:
                                return (0, f"未知映射状态: {mapping_status}")
                except Exception as e:
                    return (0, f"处理第{row}行配置时出错: {str(e)}")
            
            print(f"当前行数为：{inspect.currentframe().f_lineno}，restore_current_data成功")
            return (1, "数据恢复成功")
    
        except Exception as e:
            error_msg = f"恢复数据失败: {str(e)}"
            print(error_msg)
            return (0, error_msg)

    def One_click_clear(self):
        """一键清空配置"""
        # 清空no_mapping_tab配置表
        for row in range(0, self.table_row_number):
            # title行数初始化为1
            mapping_row_combo = self.Compare_Config_table.cellWidget(row, self.title_rows)
            mapping_row_combo.setValue(1)

            # title mapping状态改为N
            mapping_status_combo = self.Compare_Config_table.cellWidget(row, self.mapping_option)
            mapping_status_combo.setCurrentText("N")

            # workboot2的sheet name list填充到table的第一列清空
            sheet_combo = self.Compare_Config_table.cellWidget(row, 0)
            sheet_combo.clear()
            sheet_combo.addItems([""])
            sheet_combo.setCurrentIndex(0)

            # workboot2的sheet name list填充到table的第二列清空
            sheet_combo = self.Compare_Config_table.cellWidget(row, 1)
            sheet_combo.clear()
            sheet_combo.addItems([""])
            sheet_combo.setCurrentIndex(0)

        # 清空文件选择器
        self.file1_selector.set_file_path(None)
        self.file2_selector.set_file_path(None)
        self.wb1 = None
        self.wb2 = None

        # 删除config文件
        os.remove(json_file_path)

    def open_result_file(self, up_or_down):
        if up_or_down == 0:
            if not self.file_operator.open_excel_file(self.output_file_path1):
                return False
            print(f"成功打开Excel文件1: {self.output_file_path1}")
        elif up_or_down == 1:
            if not self.file_operator.open_excel_file(self.output_file_path2):
                return False
            print(f"成功打开Excel文件2: {self.output_file_path2}")
        else:
            print("未知错误")
            return False
        return True
    
    def open_log_file(self):
        if not self.file_operator.open_text_file(compare_info_file_path):
            return False
        return True

    def init_no_mapping_tab(self):
        """初始化No Mapping选项卡"""
        layout = QVBoxLayout(self.Compare_Config)# 设置大小策略

        # 设置最大高度
        self.tab_widget.setMaximumHeight(1200)
        
        # 创建表格
        self.Compare_Config_table = QTableWidget(self.table_row_number, self.table_column_number)
        self.Compare_Config_table.setHorizontalHeaderLabels([
            "文件1-sheet", "文件2-sheet", "索引列1", "索引列2", "索引列3", "mapping title", "表头行数"
        ])
        
        
        # 初始化表格行
        for row in range(self.table_row_number):
            # 设置表格行高
            self.Compare_Config_table.setRowHeight(row, self.table_row_height)
            self.init_table_row(self.Compare_Config_table, row)
        
        # 设置表格样式
        self.setup_table_style(self.Compare_Config_table)
        
        # 设置列宽
        self.Compare_Config_table.setColumnWidth(0, 120)  # 文件1 Sheet
        self.Compare_Config_table.setColumnWidth(1, 120)  # 文件2 Sheet
        self.Compare_Config_table.setColumnWidth(2, 80)   # Index Col1
        self.Compare_Config_table.setColumnWidth(3, 80)   # Index Col2
        self.Compare_Config_table.setColumnWidth(4, 80)   # Index Col3
        self.Compare_Config_table.setColumnWidth(5, 100)  # Mapping Title
        self.Compare_Config_table.setColumnWidth(6, 80)   # Title Row
        layout.addWidget(self.Compare_Config_table)

    def setup_table_style(self, table):
        """设置表格样式"""
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.verticalHeader().setVisible(False)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                alternate-background-color: #F5F5F7;
                border: 1px solid #E5E5EA;
                border-radius: 12px;
                gridline-color: #E5E5EA;
                selection-background-color: rgba(0, 122, 255, 0.15);
                selection-color: #007AFF;
                outline: none;  /* 移除选中时的虚线框 */
            }
            
            /* 表头样式 */
            QHeaderView::section {
                background-color: #F5F5F7;
                color: #333333;
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 12px;
                font-weight: 800;
                padding: 10px 15px;
                border: none;
                border-bottom: 1px solid #E5E5EA;
                border-right: 1px solid #E5E5EA;
            }
            
            QHeaderView::section:last {
                border-right: none;
            }
            
            /* 表格项样式 */
            QTableWidget::item {
                font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
                font-size: 11px;
                color: #333333;
                padding: 6px 12px;
            }
            
            /* 行高 */
            QTableWidget QTableCornerButton::section {
                background-color: #F5F5F7;
                border: none;
                border-bottom: 1px solid #E5E5EA;
                border-right: 1px solid #E5E5EA;
            }
            
            /* 滚动条样式 */
            QScrollBar:vertical {
                border: none;
                background: #F5F5F7;
                width: 12px;
                margin: 0;
                border-radius: 6px;
            }
            
            QScrollBar::handle:vertical {
                background: #C7C7CC;
                min-height: 30px;
                border-radius: 6px;
            }
            
            QScrollBar::handle:vertical:hover {
                background: #A2A2A7;
            }
            
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0;
                subcontrol-position: bottom;
                subcontrol-origin: margin;
            }
            
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            
            QScrollBar:horizontal {
                border: none;
                background: #F5F5F7;
                height: 12px;
                margin: 0;
                border-radius: 6px;
            }
            
            QScrollBar::handle:horizontal {
                background: #C7C7CC;
                min-width: 30px;
                border-radius: 6px;
            }
            
            QScrollBar::handle:horizontal:hover {
                background: #A2A2A7;
            }
            
            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                width: 0;
                subcontrol-position: right;
                subcontrol-origin: margin;
            }
            
            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: none;
            }
        """)
        # 设置表格属性
        table.setAlternatingRowColors(True)  # 启用交替行颜色
        table.setSelectionBehavior(QAbstractItemView.SelectRows)  # 整行选择
        table.setSelectionMode(QAbstractItemView.SingleSelection)  # 单选
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 自动拉伸列宽
        # table.horizontalHeader().setHighlightSections(False)  # 禁用表头高亮
        table.verticalHeader().setVisible(False)  # 隐藏垂直表头
        # 设置表格的最小和最大高度为相同值，实现固定高度
        table.setMinimumHeight(self.table_heigh)
        print("table.setFixedHeight(self.table_heigh)")

    def init_table_row(self, table, row):
        """初始化表格行"""
        table.setRowHeight(row, self.table_row_height)
        # Sheet名称下拉框
        for col in range(0, 2):
            sheet_combo = QComboBox()
            sheet_combo.setMinimumHeight(self.select_edit_heigh)  # 增加下拉框高度
            sheet_combo.addItems([""])
            sheet_combo.currentIndexChanged.connect(lambda: self.sheet_selected(table, row))
            table.setCellWidget(row, col, sheet_combo)
        
        # 索引列/数据列初始为文本框
        for col in range(self.index_col_position[0], (self.index_col_position[1]+1)):
            col_input = QLineEdit()
            col_input.setMinimumHeight(self.select_edit_heigh)  # 增加输入框高度
            col_input.setValidator(UpperCaseValidator())
            col_input.setPlaceholderText("输入大写字母")
            table.setCellWidget(row, col, col_input)
        
        # Mapping title下拉框
        mapping_combo = QComboBox()
        mapping_combo.addItems(["N", "Y"])
        mapping_combo.currentIndexChanged.connect(lambda: self.mapping_status_changed(table, row))
        table.setCellWidget(row, self.mapping_option, mapping_combo)
        
        # 表头行数
        header_spin = QSpinBox()
        header_spin.setRange(1, 500)
        header_spin.setEnabled(False)  # 初始禁用
        header_spin.valueChanged.connect(lambda: self.title_row_changed(table, row))
        table.setCellWidget(row, 6, header_spin)
        
        # 初始禁用除sheet名称外的所有控件
        for col in range(2, self.table_column_number):
            widget = table.cellWidget(row, col)
            if widget:
                widget.setEnabled(False)
    
    def table_row_number_changed(self):
        """用户修改表格行数时执行"""
        self.table_row_number = int(self.table_row_number_combo.currentText())
        
        tab_index = self.tab_widget.indexOf(self.Compare_Config)
        if tab_index != -1:
            # 删除 "Compare Config" 选项卡
            self.tab_widget.removeTab(tab_index)
        # 重新创建 Compare_Config 页面
        self.Compare_Config = QWidget()
        self.Compare_Config.setStyleSheet("""
            font-family: "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
            font-size: 11px;       /* 设置字体大小为14px */
            font-weight: 400;      /* 设置字体粗细为中等（500） */
        """)
        self.init_no_mapping_tab()
        # 添加新的 "Compare Config" 选项卡
        self.table_heigh = self.table_row_number * self.table_row_height+100  # table的高度
        self.tab_widget.setMinimumHeight(self.table_heigh)
        self.tab_widget.setMaximumHeight(self.table_heigh)
        self.tab_widget.addTab(self.Compare_Config, "Compare Config")
        # sheet列重新添加选项
        if self.file1_selector.get_file_path():
            self.add_addItems_for_combo(self.table_row_number, self.Compare_Config_table, 0, self.wb1.sheetnames)
        if self.file2_selector.get_file_path():
            self.add_addItems_for_combo(self.table_row_number, self.Compare_Config_table, 1, self.wb2.sheetnames)
        # self.restored_config_data.update_row_number(self.table_row_number)
        whole_size = self.path_edit_height*2+self.table_heigh+self.select_edit_heigh+20+self.progress_bar_heigh+self.current_task_label_height
        self.on_tab_widget_resize(whole_size)
        self.button_up.setEnabled(0)
        self.button_down.setEnabled(0)
        self.button_log.setEnabled(0)

    def on_tab_widget_resize(self, whole_size):
        # 获取当前窗口大小
        current_size = self.size()
        
        # 计算新的窗口大小（例如，根据表格的大小）
        new_height = max(current_size.height(), whole_size)
        
        # 调整窗口大小
        self.resize(QSize(current_size.width(), 10))

    def sheet_selected(self, table, row):
        """Sheet名称选择事件处理"""
        sheet1 = table.cellWidget(row, 0).currentText()
        sheet2 = table.cellWidget(row, 1).currentText()
        # 当两个sheet都选择后启用其他控件
        if sheet1 and sheet2:
            for col in range(2, self.table_column_number):
                widget = table.cellWidget(row, col)
                if widget:
                    widget.setEnabled(True)
            
            # 触发mapping状态更新
            self.mapping_status_changed(table, row)
        else:
            for col in range(2, self.table_column_number):
                widget = table.cellWidget(row, col)
                if widget:
                    widget.setEnabled(False)
        self.button_up.setEnabled(0)
        self.button_down.setEnabled(0)
        self.button_log.setEnabled(0)
        
    def title_row_changed(self, table, row):
        """title列数变更-事件处理"""
        sheet1_name = table.cellWidget(row, 0).currentText()
        sheet2_name = table.cellWidget(row, 1).currentText()
        title_rows_number = table.cellWidget(row, self.title_rows).value()
        self.title_list = self.get_title_list(self.wb1[sheet1_name], self.wb2[sheet2_name], title_rows_number)
        print(f"self.title_list = {self.title_list}")
        # string_list_model = QStringListModel(self.title_list)
        for col in range(self.index_col_position[0], self.index_col_position[1]+1):
            combo = table.cellWidget(row, col)
            combo.clear()
            combo.addItems([""]+self.title_list)
            combo.completer().setModel(QStringListModel(self.title_list))
        
        self.button_up.setEnabled(0)
        self.button_down.setEnabled(0)
        self.button_log.setEnabled(0)
        
    def get_title_list(self, sheet1, sheet2, title_row_number):
        title_row_values1 = list(next(sheet1.iter_rows(min_row=title_row_number, max_row=title_row_number, values_only=True)))
        title_row_values2 = list(next(sheet2.iter_rows(min_row=title_row_number, max_row=title_row_number, values_only=True)))
        title_list = []
        for value1 in title_row_values1:
            for value2 in title_row_values2:
                # print(f"value1 = {value1}, value2 = {value2}")
                if str(value1) == str(value2) and str(value1) and value1 != None:
                    title_list.append(str(value1))
                    break

        print(f"title_list = {title_list}")
        return title_list
    
    def mapping_status_changed(self, table, row):
        """Mapping title选择事件处理"""
        mapping_combo = table.cellWidget(row, self.mapping_option)
        mapping_value = mapping_combo.currentText()
        header_spin = table.cellWidget(row, self.title_rows)
        
        if mapping_value == 'Y':
            # 启用表头行数输入
            header_spin.setEnabled(True)
            # 获取index列的选项卡list
            sheet1_name = table.cellWidget(row, 0).currentText()
            sheet2_name = table.cellWidget(row, 1).currentText()
            print(f"sheet1_name ={sheet1_name}")
            print(f"sheet2_name ={sheet2_name}")
            title_rows_number = header_spin.value()
            # 基于第index个索引框内容，查找两个sheet中一致的title名称list
            self.title_list = self.get_title_list(self.wb1[sheet1_name], self.wb2[sheet2_name], title_rows_number)
            
            # 将索引列/数据列转换为带搜索功能的下拉框
            for col in range(self.index_col_position[0], (self.index_col_position[1]+1)):
                # 移除现有控件
                old_widget = table.cellWidget(row, col)
                if old_widget:
                    table.removeCellWidget(row, col)
                
                # 创建下拉框 - 可编辑，带搜索功能
                combo = QComboBox()
                combo.setEditable(True)
                combo.addItem("")  # 添加空选项

                combo.addItems(self.title_list) #将所有title放入index的选项中
                combo.setCurrentIndex(0)  # 默认选择空项
                
                # 设置自动补全功能
                completer = QCompleter(self.title_list)
                completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)  # 不区分大小写
                completer.setFilterMode(Qt.MatchFlag.MatchContains)  # 包含匹配
                combo.setCompleter(completer)
                
                # 设置只能选择列表中的值
                def validate_input(combo=combo):
                    text = combo.currentText()
                    if text not in self.title_list:
                        # 如果输入不在选项中，重置为之前的值
                        index = combo.findText(text, Qt.MatchFlag.MatchExactly)
                        if index == -1:
                            combo.setCurrentIndex(0)  # 重置为空选项
                        else:
                            combo.setCurrentIndex(index)
                
                # 连接编辑完成信号
                combo.lineEdit().editingFinished.connect(validate_input)
                table.setCellWidget(row, col, combo)
        else:
            # 禁用表头行数并重置
            header_spin.setEnabled(False)
            header_spin.setValue(1)
            
            # 将下拉框转换为文本框
            for col in range(self.index_col_position[0], (self.index_col_position[1]+1)):
                # 移除现有控件
                old_widget = table.cellWidget(row, col)
                if old_widget:
                    table.removeCellWidget(row, col)
                
                # 创建文本框
                line_edit = QLineEdit()
                line_edit.setValidator(UpperCaseValidator())
                line_edit.setPlaceholderText("输入大写字母")
                table.setCellWidget(row, col, line_edit)
        
        self.button_up.setEnabled(0)
        self.button_down.setEnabled(0)
        self.button_log.setEnabled(0)

    def on_comparison_finished(self, complete_flag):   #线程完成的槽函数，当进度达到100%时或用户强制终止时发射信号调用
        self.set_button_status("开始处理")
        self.button_up.setEnabled(complete_flag == "success")
        self.button_down.setEnabled(complete_flag == "success")
        self.button_log.setEnabled(complete_flag == "success")
        print(f"bool(complete_flag) = {(complete_flag)}")

    def add_addItems_for_combo(self, row_number, table, column, Option_value_list):
        """给table的列添加选项值"""
        for row in range(0, row_number):
            sheet_combo = table.cellWidget(row, column)
            sheet_combo.clear()
            sheet_combo.addItems([""] + Option_value_list)
    
    def get_file_output_path_byFilepath(self, file_path):
        # 处理文件路径和输出路径
        file_name, file_ext = os.path.splitext(os.path.basename(file_path))
        # 处理输出文件扩展名：CSV文件默认转为XLSX，其他格式保留原扩展名
        if file_ext.lower() == '.csv':
            output_ext = '.xlsx'
        else:
            output_ext = file_ext  # 保留原扩展名（如.xlsx、.xlsm等）

        # 构建完整输出路径
        outputfile_path = os.path.join(output_path, f"{file_name}-compare{output_ext}")
        return outputfile_path

    def browse_file(self, selector):
        """浏览文件对话框"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择文件",
            "",
            "Excel文件 (*.xlsx *.xls *.xlsm *.csv)"
        )
        
        if file_path:
            try:
                selector.set_file_path(file_path)
                wb, error_msg = DataProcessor.open_file(file_path, True)
                print("ValueError(error_msg)1")
                if wb is None:
                    ValueError(error_msg)
                    print("ValueError(error_msg)2")
                print("ValueError(error_msg)3")
                print(f"wb.sheetnames = {wb.sheetnames}")
                if selector == self.file1_selector:
                    self.wb1 = wb
                    self.add_addItems_for_combo(self.table_row_number, self.Compare_Config_table, 0, self.wb1.sheetnames)
                    # 处理文件路径和输出路径
                    self.output_file_path1 = self.get_file_output_path_byFilepath(file_path)
                elif selector == self.file2_selector:
                    self.wb2 = wb
                    self.add_addItems_for_combo(self.table_row_number, self.Compare_Config_table, 1, self.wb2.sheetnames)
                    self.output_file_path2 = self.get_file_output_path_byFilepath(file_path)
                else:
                    pass
            except Exception as e:
                self.current_task_edit.appendPlainText(f"无法打开文件: {file_path}\n{str(e)}")
                print(f"无法打开文件: {file_path}\n{str(e)}")
                QMessageBox.warning(self, "错误", f"无法打开文件: {file_path}\n{str(e)}")
                return 0
        
        self.button_up.setEnabled(0)
        self.button_down.setEnabled(0)
        self.button_log.setEnabled(0)

    def check_files_selected(self):
        """检查文件是否都已选择"""
        file1_selected = bool(self.file1_selector.get_file_path())
        file2_selected = bool(self.file2_selector.get_file_path())
        self.start_btn.setEnabled(file1_selected and file2_selected)


    def start_comparison(self):
        """开始/停止对比处理"""
        # 创建线程来执行文件打开和对比操作
        if self.start_flag:
            if self.start_processing():
                self.set_button_status("停止处理")
                self.progress_bar.setValue(0)
                self.button_up.setEnabled(0)
                self.button_down.setEnabled(0)
                self.button_log.setEnabled(0)
                self.current_task_edit.clear()
            else:
                pass
        else:
            self.processor.stop()
            self.processor.wait()  # 等待线程结束（可选）
            self.set_button_status("开始处理")

    def set_button_status(self, status):
        """设置按钮状态"""
        if status == "停止处理":
            self.start_flag = False
            self.start_btn.setEnabled(False)
            self.start_btn.setText("停止处理")
            self.start_btn.setEnabled(True)
            print(f"当前行数为：{inspect.currentframe().f_lineno}, Button set to Stop Button.")
        elif status == "开始处理":
            self.start_flag = True
            self.start_btn.setEnabled(False)
            self.start_btn.setText("开始处理")
            self.start_btn.setEnabled(True)
            print(f"当前行数为：{inspect.currentframe().f_lineno}, set to Start Button.")
        else:
            print(f"当前行数为：{inspect.currentframe().f_lineno}, set_button_status status input error, Start or Stop?")
            return 0
        return 1

    def start_processing(self):
        """开始处理数据"""
        current_tab = self.tab_widget.currentWidget()
        current_table = None
        if current_tab == self.Compare_Config:
            current_table = self.Compare_Config_table
        
        if not current_table:
            QMessageBox.warning(self, "错误", "无法确定当前配置表格")
            return 0
        print(f"当前行数为：{inspect.currentframe().f_lineno} start_processing")
        
        # 收集配置数据
        config_data = []
        for row in range(self.table_row_number):
            row_data = []
            for col in range(self.table_column_number):
                widget = current_table.cellWidget(row, col)
                if widget:
                    if isinstance(widget, QComboBox):
                        row_data.append(widget.currentText())
                    elif isinstance(widget, QLineEdit):
                        row_data.append(widget.text())
                    elif isinstance(widget, QSpinBox):
                        row_data.append(widget.value())
                else:
                    row_data.append("")
            config_data.append(row_data)
        print(f"当前行数为：{inspect.currentframe().f_lineno} start_processing")
        
        # 验证必要字段
        valid = False
        unique_values1 = []
        unique_values2 = []
        for row in config_data:
            if row[0] and row[1]:  # 两个sheet名称都有值
                unique_values1.append(row[0])
                unique_values2.append(row[1])
                if row[self.mapping_option] == 'Y':  # mapping模式
                    # 检查是否选择了有效的选项（非空）
                    if any(row[self.index_col_position[0]:(self.index_col_position[1]+1)]) and row[self.title_rows] > 0:
                        valid = True
                    else:
                        valid = False
                        break
                else:  # no mapping模式
                    # if any(row[2:5]):
                    #     valid = True
                    valid = True
        counter1 = Counter(unique_values1)
        duplicates1 = {element: count for element, count in counter1.items() if count > 1}
        counter2 = Counter(unique_values2)
        duplicates2 = {element: count for element, count in counter2.items() if count > 1}
        if duplicates1:
            QMessageBox.warning(self, "配置错误", f"文件1-sheet列中存在重复元素 {duplicates1} ")
            return 0
        if duplicates2:
            QMessageBox.warning(self, "配置错误", f"文件2-sheet列中存在重复元素 {duplicates2} ")
            return 0
        print(f"config_data = {config_data}")
        if not valid:
            QMessageBox.warning(
                self, 
                "配置错误", 
                "请确保至少一行配置完整：\n"
                "- 选择两个sheet名称\n"
                "- 至少一个索引列/数据列\n"
                "- 如果是mapping模式，表头行数需大于0"
            )
            return 0
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.result_text_edit.setText("处理中...")
        self.start_btn.setEnabled(False)
        
        # 创建并启动处理线程
        self.processor = DataProcessor(
            self.file1_selector.get_file_path(),
            self.file2_selector.get_file_path(),
            config_data
        )
        
        self.processor.progress_updated.connect(self.update_progress)
        self.processor.result_text_edit.connect(self.show_result)
        self.processor.error_occurred.connect(self.show_error)
        self.processor.comparison_finished.connect(self.on_comparison_finished)
        self.processor.progress_current_task.connect(self.current_task_edit.appendPlainText)
        
        self.processor.finished.connect(self.processing_finished)
        self.processor.start()
        return 1

    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar.setValue(value)

    def show_result(self, result):
        """显示处理结果"""
        self.result_text_edit.setText(result)

    def show_error(self, error):
        """显示错误信息"""
        self.result_text_edit.setHtml(f"<font color='red'>{error}</font>")
        QMessageBox.critical(self, "处理错误", error)

    def processing_finished(self):
        """处理完成后的清理工作"""
        self.start_btn.setEnabled(True)
    
    def get_combo_all_options(self, sheet_combo):
        # 获取所有选项
        all_items = []
        for i in range(sheet_combo.count()):
            item_text = sheet_combo.itemText(i)
            all_items.append(item_text)
        return all_items

def setup_window_geometry(window, window_width, window_height):
    """将窗口居中显示（PySide6 兼容版）"""
    # 获取主屏幕
    screen = QGuiApplication.primaryScreen()
    screen_geometry = screen.geometry()  # 屏幕尺寸
    
    # 计算居中坐标（水平方向距左400px，垂直方向居中）
    x = (screen_geometry.width() - window_width) // 2 # 水平方向固定偏移
    y = 100
    
    # 应用窗口位置和大小
    window.setGeometry(x, y, window_width, window_height)

def resource_path(relative_path):
    """获取资源的绝对路径（处理打包前后的路径差异）"""
    if getattr(sys, 'frozen', False):  # 判断程序是否被打包
        # Nuitka 打包后，sys.executable 指向 .exe 文件
        base_path = os.path.dirname(sys.executable)
    else:
        # 开发阶段，使用当前脚本所在目录
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    # 拼接资源路径（注意使用 os.path.join 确保跨平台兼容）
    return os.path.join(base_path, relative_path)  

class InitialScreen(QWidget):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.setup_ui()
        self.main_window = main_window
        
        # 加载背景图片（使用修复后的图片路径）
        self.background_pixmap = QPixmap(resource_path("ICO/GUI_ICO2.png"))
        if self.background_pixmap.isNull():
            # 尝试加载原始图片（如果修复后的图片不存在）
            self.background_pixmap = QPixmap(resource_path("ICO/GUI_ICO2.png"))
            if self.background_pixmap.isNull():
                print("警告：背景图片加载失败")
    
    def setup_ui(self):
        # 设置窗口标题和大小
        self.setWindowTitle("欢迎使用")
        self.resize(800, 600)
        
        # 创建垂直布局
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)  # 去除边距
        layout.setSpacing(0)
        
        # 创建一个用于放置按钮的容器（透明）
        button_container = QWidget()
        button_container.setStyleSheet("background-color: transparent;")
        button_layout = QVBoxLayout(button_container)
        button_layout.setAlignment(Qt.AlignBottom | Qt.AlignHCenter)  # 底部中央
        button_layout.setContentsMargins(0, 0, 0, 100)  # 底部边距
        
        # 创建开始按钮
        start_button = QPushButton("开始探索")
        start_button.setMinimumSize(200, 60)  # 增大按钮尺寸，更符合苹果风格
        start_button.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgba(148, 217, 255, 0.95),
                                                stop:1 #007AFF);
                        color: white;
                        border-radius: 16px;
                        font-family: "SF Pro Text";
                        font-size: 25px;
                        font-weight: 800;
                        padding: 16px 32px;
                        border: none;
                        box-shadow: 0 4px 12px rgba(0, 122, 255, 0.2);
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgba(180, 228, 255, 0.98),
                                                stop:1 #0066CC);
                        box-shadow: 0 8px 16px rgba(0, 122, 255, 0.3);
                    }
                    QPushButton:pressed {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgba(100, 180, 255, 0.9),
                                                stop:1 #005299);
                        box-shadow: 0 2px 6px rgba(0, 122, 255, 0.1);
                    }
        """)
        # 将按钮添加到容器布局
        button_layout.addWidget(start_button)
        
        # 将按钮容器添加到主布局
        layout.addWidget(button_container)
        
        # 设置按钮点击事件
        start_button.clicked.connect(self.on_start_clicked)
    
    def paintEvent(self, event):
        """重写绘制事件，绘制背景图片（避免样式表中的图片路径问题）"""
        if not self.background_pixmap.isNull():
            painter = QPainter(self)
            # 缩放图片以覆盖整个窗口（保持宽高比，可能裁剪）
            scaled_pixmap = self.background_pixmap.scaled(
                self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation
            )
            # 居中绘制图片
            x = (self.width() - scaled_pixmap.width()) // 2
            y = (self.height() - scaled_pixmap.height()) // 2
            painter.drawPixmap(x, y, scaled_pixmap)
        else:
            # 图片加载失败时，绘制灰色背景
            painter = QPainter(self)
            painter.fillRect(self.rect(), Qt.gray)
            # 添加错误文本提示
            error_label = QLabel("背景图片加载失败", self)
            error_label.setAlignment(Qt.AlignCenter)
            error_label.setStyleSheet("color: white; font-size: 16px;")
            error_label.setGeometry(0, 0, self.width(), self.height())
    
    def license_verify(self, license_file_path):
        print("开始许可证验证")
        verify_app = DeviceIDLicenseVerify(license_file_path)
        if not verify_app.verify_license():
            print("=" * 40)
            print("授权验证失败，程序无法继续运行。")
            print("请联系开发者获取有效的授权文件。")
            sys.exit(1)
            return 0
        
        # 授权通过，运行主程序
        print("=" * 40)
        print("Hello World!")
        print("这是一个经过授权的程序。")
        print("=" * 40)
        return 1
    
    def on_start_clicked(self):
        global license_file_path
        # 点击按钮后，显示主界面并关闭当前界面
        # 验证签名
        if self.license_verify(license_file_path):
            # 打开主窗口
            self.main_window.show()
        
        # 关闭初始化窗口
        self.close()

        

"""程序入口点：初始化应用程序并启动事件循环"""
try:

    # 初始化 Qt 应用程序实例，处理命令行参数
    app = QApplication(sys.argv)

    # 设置应用程序样式为 Fusion（跨平台统一风格，更现代）
    app.setStyle("Fusion")

    
    
    # 设置全局字体为 Segoe UI（Windows 默认字体），大小为 10 磅
    # 确保界面文字在不同系统上显示一致
    font = QFont("Segoe UI", 10)
    app.setFont(font)

    # 创建主窗口实例（DataProcessingTool 类应继承自 QMainWindow 或 QWidget）
    window = DataProcessingTool()
    initial_screen = InitialScreen(window)
    initial_screen.show()

    # 进入 Qt 应用程序的事件循环，等待用户交互或系统事件
    # sys.exit() 确保应用程序退出时返回正确的状态码
    sys.exit(app.exec())

except Exception as e:
    # 捕获应用程序启动过程中的任何异常（如导入错误、初始化失败）
    # 将异常信息写入错误日志文件（即使程序崩溃也能追踪问题）
    with open('error.log', 'w', encoding='utf-8') as f:
        f.write(f'错误：{str(e)}\n')
        f.write(traceback.format_exc())