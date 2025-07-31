# -*- coding: utf-8 -*-
## 基于DMEO3修改
## 增加.xlsm格式支持，添加Person_ComparisonApp类
## 增加当前task显示框及逻辑
## 文件对比功能基本修改完毕，
## 待办：保存文件步骤，用于处理结束的槽函数编写
## 待办：设置sheet或index表头列的选项list，选择文件后能够直接选择文件中的sheet和sheet中的title
## 待办：增加一列用于存放title所在的行
## 待办：不mapping title时，index能够分别选择两个文件的index_col，以解决两个文件中的index列不在同一列的问题 --（还是拆分两个表格会好一点，不会混乱，如果保持现状，可能需要在修改mapping title时重建表格，需调研一下）
## 
## 待办：增加当前task框，放入当前进度，停止按钮还是失效，需要修复
## 已修复
## 待办：确认mapping row和mapping col的性能优化效果，继续优化大文件的对比速度
## 2025/07/08：修复索引列空白时，报“用户手动终止进程”的问题
## 2025/07/29：V16,对比结果框中添加"对比结果摘要"，同时将摘要放入对比结果log文件中，以快速浏览对比结果
##


import inspect
import copy
import json
import ctypes
import time
import threading
import openpyxl, xlrd
import textwrap
from collections import Counter
from typing import List, Dict, Set, Tuple, Optional, Union
from collections import defaultdict
import sys
import os
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTabWidget,
    QTableWidget, QTableWidgetItem, QComboBox, QLineEdit, QPushButton, QLabel,
    QFileDialog, QMessageBox, QProgressBar, QHeaderView, QSpinBox, QCompleter,
)
from PySide6.QtCore import Qt, QThread, Signal, QStringListModel
from PySide6.QtGui import QColor, QFont, QValidator

class Person_ComparisonApp:
    is_running = True   #调用function终止本类函数的运行，例如：Person_ComparisonApp.is_running = False即可终止
    # 类的构造函数，用于初始化对象的属性
    def __init__(self, progress_updated, progress_current_task, comparison_finished, output_path=None):
        """
        Excel对比应用核心类，负责执行表格对比逻辑
        
        Args:
            progress_updated: 进度更新信号，用于更新UI进度条
            progress_current_task: 任务进度信号，用于显示当前处理任务
            comparison_finished: 对比完成信号，用于通知UI对比结束
            output_path: 输出文件路径（可选）
        """
        self.progress_updated       = progress_updated      #用于更新当前进度的信号槽，比如：self.progress_updated(xxx)返回当前进度%
        self.progress_current_task  = progress_current_task #用于更新当前正在进行的task名称，比如：第x行正在对比
        self.comparison_finished    = comparison_finished   #用于返回当前进程已终止，原因：“被用户手动停止”
        self.output_path = output_path                    # 输出文件路径
        self.Progress_percent = 0                         # 当前进度百分比
        self.Agreed_color = "AFFFAF"  # 一致时填充色（浅绿色）
        self.Not_Agreed_color = "C04255"  # 不一致时填充色（红色）
        self.No_match_color = "00FFFF"  # 未匹配时填充色（青色）
        self.None_color = "FFFFFF"  # 空值填充色（白色）
        self.Delete_color = "1F1F1F"  # 删除行填充色（灰色）
        self.update_frequency = 20  # 状态更新频率，防止UI卡顿
        self.result_info = None     #用于保存本次对比结果


    def check_index_repeat(self, sheet, index_value_list: List[int], title_row_number: int=0, file_name:str="") -> int:
        """
        检查工作表中索引列是否存在重复值
        
        Args:
            sheet: 当前工作表对象
            index_value_list: 索引列的列号列表
            title_row_number: 标题行号（默认为0，表示无标题行）
            file_name: 文件名（用于错误提示）
            
        Returns:
            1: 检查完成
            0: 任务被终止
        """
        # 过滤无效索引列（值为0的列）
        index_columns = [col for col in index_value_list if col != 0]
        if not index_columns:
            return 1  # 没有有效的索引列，直接返回
        
        merged_values: Dict[str, int] = {}  # 记录已出现的合并值及其首次出现的行号
        blank_row_count = 0
        max_row = sheet.max_row
        for row_num in range(title_row_number+1, max_row + 1):
            # 检查任务是否被终止
            if self.check_thread_running():
                return 0
                
            # 获取当前行的合并文本
            merged_text = self.get_merged_text(sheet, row_num, index_columns)
            
            # 更新进度
            if row_num % self.update_frequency == 0:
                self.progress_current_task.emit(
                    f"正在检查sheet【{sheet.title}】中索引列索引列是否有重复元素，检查第{row_num}/{max_row}行"
                )
            
            # 处理空行检测
            if not merged_text:
                blank_row_count += 1
                if blank_row_count >= 20:
                    print_info = f"连续20行索引列值为空，结束sheet【{sheet.title}】的检查"
                    print(print_info)
                    self.progress_current_task.emit(print_info)
                    break
                continue
            else:
                blank_row_count = 0
            # print(f"当前行数为：{inspect.currentframe().f_lineno}，check_index_repeat")
            
            # 检查重复值
            if merged_text in merged_values:
                first_row = merged_values[merged_text]
                error = (
                    f"文档中\nsheet：[{sheet.title}]\n"
                    f"第[{index_columns}]列合并后存在相同元素\n"
                    f"第{first_row}行与第{row_num}行相同：【{merged_text}】"
                )
                self.progress_current_task.emit(error)
                raise ValueError(error)
            
            # 记录新出现的合并值
            merged_values[merged_text] = row_num
        
        return 1
    
    # 拆分合并单元格，并复制格式信息。
    def get_merged_text(self, sheet, row, index_value_list):
        """
        获取指定行中索引列的合并文本（用于索引匹配）
        
        Args:
            sheet: 工作表对象
            row: 行号
            index_value_list: 索引列号列表
            
        Returns:
            str: 合并后的索引文本
        """
        merged_text = ""
        for col in index_value_list:
            value = sheet.cell(row, col).value
            if value is not None:
                # 提前处理字符串，避免后续重复替换
                value = self._process_title_text(str(value))
                merged_text += value
            
        return merged_text

    def copy_cell_format(self, sheet, row1, col1, row_min, col_min):
        """
        复制单元格格式（填充、边框、字体等）
        
        Args:
            sheet: 工作表对象
            row1, col1: 目标单元格坐标
            row_min, col_min: 源单元格坐标（左上角合并单元格）
        """
        # 复制填充格式
        if sheet.cell(row=row1, column=col1).fill:
            start_color = sheet.cell(row=row_min, column=col_min).fill.start_color
            end_color = sheet.cell(row=row_min, column=col_min).fill.end_color
            fill_type = sheet.cell(row=row_min, column=col_min).fill.fill_type
            sheet.cell(row=row1, column=col1).fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
        
        # 复制边框格式
        if sheet.cell(row=row1, column=col1).border:
            left_border = sheet.cell(row=row_min, column=col_min).border.left
            right_border = sheet.cell(row=row_min, column=col_min).border.left
            top_border = sheet.cell(row=row_min, column=col_min).border.top
            bottom_border = sheet.cell(row=row_min, column=col_min).border.bottom
            sheet.cell(row=row1, column=col1).border = openpyxl.styles.Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
        
        # 复制字体格式
        if sheet.cell(row=row_min, column=col_min).font:
            font = sheet.cell(row=row_min, column=col_min).font
            sheet.cell(row=row1, column=col1).font = openpyxl.styles.Font(
                name=font.name, size=font.size, color=font.color,
                bold=font.bold, italic=font.italic, underline=font.underline
            )
        
        # 复制对齐格式
        align = sheet.cell(row=row_min, column=col_min).alignment
        sheet.cell(row=row1, column=col1).alignment = openpyxl.styles.Alignment(
            horizontal=align.horizontal, vertical=align.vertical, wrap_text=align.wrap_text
        )
        
        # 复制数字格式
        sheet.cell(row=row1, column=col1).number_format = sheet.cell(row=row_min, column=col_min).number_format

    def split_merged_cells(self, sheet, index_col_list = []):
        """
        拆分合并单元格并复制格式
        
        Args:
            sheet: 工作表对象
            index_col_list: 索引列列表（用于特殊处理索引列）
            
        Returns:
            int: 1表示成功，0表示任务终止
        """
        self.progress_current_task.emit(f"开始拆分")
        merged_cell_ranges = list(sheet.merged_cells.ranges)
        for merged_cell in merged_cell_ranges:
            self.progress_current_task.emit(f"merged_cell = {merged_cell}")
            row_min = merged_cell.min_row
            row_max = merged_cell.max_row
            col_min = merged_cell.min_col
            col_max = merged_cell.max_col
            sheet.unmerge_cells(start_row=row_min, start_column=col_min, end_row=row_max, end_column=col_max)
            for row1 in range(row_min, row_max+1):
                if self.check_thread_running():
                    return 0
                for col1 in range(col_min, col_max+1):
                    if self.check_thread_running():
                        return 0
                    # print(f"当前行数为：{inspect.currentframe().f_lineno} row = {row1}")
                    if not(row1 == row_min and col1 == col_min):
                        # 计算相对位置
                        index_row = row1 - row_min
                        index_col = col1 - col_min
                        sheet.cell(row=row1, column=col1).value = sheet.cell(row=row_min, column=col_min).value
                        # 判断是否是索引列
                        if col1 in index_col_list: # 单元格在索引列，赋值必须唯一{左上角值+行值+列值}
                            sheet.cell(row=row1, column=col1).value = f"{sheet.cell(row=row1, column=col1).value}{index_row}{index_col}"
                        else: # 单元格不在索引列，赋值统一为左上角值
                            sheet.cell(row=row1, column=col1).value = f"{sheet.cell(row=row1, column=col1).value}"
                        # 复制单元格格式
                        self.copy_cell_format(sheet, row1, col1, row_min, col_min)
        return 1
    
    def open_file(self, file_path):
        # 加载一个 Excel 文件
        try:
            if file_path.lower().endswith('.xls'):
                # 处理 .xls 文件
                wb = openpyxl.Workbook()
                xls_wb = xlrd.open_workbook(file_path, read_only=True)
                for sheet_name in xls_wb.sheet_names():
                    xls_sheet = xls_wb.sheet_by_name(sheet_name)
                    new_sheet = wb.create_sheet(sheet_name)
                    for row in range(xls_sheet.nrows):
                        for col in range(xls_sheet.ncols):
                            new_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)
                del wb['Sheet']  # 删除默认创建的工作表
            elif file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
                df.to_excel('data.xlsx', index=False)
                wb = openpyxl.load_workbook('data.xlsx', read_only=True)
                os.remove('data.xlsx')

                pass
            else:
                # 处理 .xlsx 和 .xlsm 文件
                wb = openpyxl.load_workbook(file_path, read_only=True)
        except FileNotFoundError:
            error = f"文件 {file_path} 不存在。"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
        except openpyxl.utils.exceptions.InvalidFileException:
            error = f"文件 {file_path} 不是有效的 Excel 文件, 请重新输入"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
        except Exception as e:
            error = f"发生了未知错误：{e}"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
        return wb
    
    def cell_consistency_check(self, sheet1_cell, sheet2_cell):
        """
        检查两个单元格的值是否一致
        
        Args:
            sheet1_cell: 第一个工作表单元格
            sheet2_cell: 第二个工作表单元格
            
        Returns:
            bool: True表示一致，False表示不一致
        """
        # 设置单元格自动换行
        sheet1_cell.alignment = Alignment(wrap_text=True)    #把第一个文件的单元格设为自动换行
        sheet2_cell.alignment = Alignment(wrap_text=True)   #把第二个文件的单元格设为自动换行
        
        # 处理单元格值（移除特殊字符）
        value1 = self._process_title_text(str(sheet1_cell.value))  #sheet1中对应单元格的值
        value2 = self._process_title_text(str(sheet2_cell.value))  #sheet2中对应单元格的值
        
        # 空值处理
        if str(value1) == "None":
            value1 = ""
        if str(value2) == "None":
            value2 = ""
        
        # 字符串对比
        if value1 == value2:  # 单元格值对比相同
            return True
        else:
            # 尝试数值对比（处理数字格式差异）
            try:
                # 尝试将单元格的值转换为浮点数
                value3 = float(sheet1_cell.value)
                value4 = float(sheet2_cell.value)
                # 比较两个浮点数
                if value3 == value4:
                    return True
                else:
                    return False
            except (ValueError, TypeError):
                # 如果无法转换为浮点数
                return False
    def clear_all_conditional_formatting(self, sheet):
        """清除工作表中的所有条件格式"""
        sheet.conditional_formatting = []
        return sheet

    def compare_excel_sheet(
        self, 
        sheet1, 
        sheet2, 
        current_progress: int = 0, 
        target_progress: int = 100
    ) -> Union[object, int]:
        """
        直接对比两个工作表（不使用索引或标题映射）
        
        Args:
            sheet1: 第一个工作表
            sheet2: 第二个工作表
            current_progress: 当前进度百分比
            target_progress: 目标进度百分比
            
        Returns:
            sheet1: 对比后的工作表
            0: 任务被终止
        """
        self.result_info = "" #清空上次对比结果信息
        print(f"直接对比: sheet_name = {sheet1.title}")
        self.clear_all_conditional_formatting(sheet1)
        row_changed_list = {}
        
        # 基础参数获取
        max_row1 = sheet1.max_row
        max_col1 = sheet1.max_column
        
        # 进度计算
        delta_progress = target_progress - current_progress
        progress_percent = current_progress
        step_progress = delta_progress / max_row1 if max_row1 > 0 else 0
        
        # 合并单元格处理
        self.progress_current_task.emit(f"正在拆分sheet【{sheet1.title}】的合并单元格")
        if not self.split_merged_cells(sheet1):
            return 0
        if not self.split_merged_cells(sheet2):
            return 0
        
        # 对比参数初始化
        blank_row_flag = 0
        
        for row1 in range(1, max_row1 + 1):
            # 进度更新
            progress_percent += step_progress
            if self.check_thread_running():
                return 0
                
            # 定期更新进度，避免UI卡顿
            if row1 % self.update_frequency == 0:
                self.progress_updated.emit(round(progress_percent))
                self.progress_current_task.emit(
                    f"正在对比sheet【{sheet1.title}】第{row1}/{max_row1}行, 进度: {progress_percent:.1f}%"
                )
            
            blank_col_flag = 0
            all_cells_empty = True
            row_isChanged_status_flag = False
            
            # 逐列对比
            for col in range(1, max_col1 + 1):
                if self.check_thread_running():
                    return 0
                    
                sheet1_cell = sheet1.cell(row=row1, column=col)
                sheet2_cell = sheet2.cell(row=row1, column=col)
                
                # 单元格一致性检查
                if self.cell_consistency_check(sheet1_cell, sheet2_cell):
                    sheet1_cell.fill = PatternFill(
                        start_color=self.Agreed_color, 
                        end_color=self.Agreed_color, 
                        fill_type="solid"
                    )
                    
                    # 空值处理
                    value1 = sheet1_cell.value or ""
                    value2 = sheet2_cell.value or ""
                    
                    if value1 == "" and value2 == "":
                        blank_col_flag += 1
                    else:
                        all_cells_empty = False
                else:
                    row_isChanged_status_flag = True
                    sheet1_cell.fill = PatternFill(
                        start_color=self.Not_Agreed_color, 
                        end_color=self.Not_Agreed_color, 
                        fill_type="solid"
                    )
                    all_cells_empty = False
            if row_isChanged_status_flag:
                row_changed_list[row1] = 2  #表征本行是否变更的flag，2：对比不一致，红色
            else:
                row_changed_list[row1] = 1  #表征本行是否变更的flag，1：对比一致，绿色
            
            # 空行检测
            if all_cells_empty:
                blank_row_flag += 1
                if blank_row_flag >= 20:
                    print_info = f"连续20行全部值为空，结束sheet【{sheet1.title}】的对比"
                    print(print_info)
                    self.progress_current_task.emit(print_info)
                    break
            else:
                blank_row_flag = 0
        
        # 创建变更标记列
        self.create_row_changed_rows(sheet1, row_changed_list)
        self.result_info += textwrap.dedent(f"""
        参与对比行数:{len(row_changed_list)}
        变更行数:{sum(1 for v in row_changed_list.values() if v == 2)}
        """)
        self.progress_updated.emit(target_progress)
        return sheet1
    
    def compare_excel_sheet_by_index(
        self, 
        sheet1, 
        sheet2, 
        index_value_list: list, 
        file1_name: str,
        current_progress: int = 0, 
        target_progress: int = 100
    ) -> Union[object, int]:
        """
        基于索引列对比两个工作表
        
        Args:
            sheet1: 第一个工作表
            sheet2: 第二个工作表
            index_value_list: 索引列号列表
            file1_name: 文件名（用于日志）
            current_progress: 当前进度百分比
            target_progress: 目标进度百分比
            
        Returns:
            sheet1: 对比后的工作表
            0: 任务被终止
        """
        self.result_info = "" #清空上次对比结果信息
        # 初始化和数据预处理
        self.clear_all_conditional_formatting(sheet1)
        start_time = time.time()
        row_changed_list = {}
        """row_changed_list用于表示本行是否存在变更，
            0: 表示index未匹配上（无颜色填充），
            1: 表示行无变更（绿色填充），
            2: 表示行有变更（红色填充），
            3: 表示行新增（青色填充），
        """
        
        max_row1 = sheet1.max_row
        max_row2 = sheet2.max_row
        max_col1 = sheet1.max_column
        
        # 索引列去重并保留顺序
        index_value_list = list(dict.fromkeys(index_value_list))
        if not index_value_list:
            return sheet1  # 没有有效索引列，直接返回
        
        # 进度计算
        delta_progress = target_progress - current_progress
        progress_percent = current_progress
        
        # --------------------- 索引列检查 --------------------- #
        self.progress_current_task.emit(f"正在检查sheet【{sheet1.title}】索引列重复值")
        if self.check_index_repeat(sheet1, index_value_list) == 0:
            return 0
        if self.check_index_repeat(sheet2, index_value_list) == 0:
            return 0
        
        check_index_time = time.time()
        check_index_time_output = f"索引检查耗时: {check_index_time - start_time}s"
        print(check_index_time_output)
        
        # --------------------- 合并单元格处理 --------------------- #
        self.progress_current_task.emit(f"正在拆分sheet【{sheet1.title}】合并单元格")
        if not self.split_merged_cells(sheet1, index_value_list):
            return 0
        if not self.split_merged_cells(sheet2, index_value_list):
            return 0
        
        split_time = time.time()
        split_time_output = f"合并单元格拆分处理耗时: {split_time - check_index_time}s"
        print(split_time_output)
        
        # --------------------- 行映射建立 --------------------- #
        self.progress_current_task.emit(f"正在建立sheet【{sheet1.title}】行映射关系")
        index_column_mapping = self.mapping_row_by_index(
            sheet1, 
            sheet2, 
            index_value_list, 
            index_value_list
        )
        if index_column_mapping == 0:
            return 0
        
        step_progress = delta_progress / len(index_column_mapping) if len(index_column_mapping) > 0 else 0

        row_mapping_time = time.time()
        row_mapping_time_output = f"行映射建立耗时: {row_mapping_time - split_time}s"
        print(row_mapping_time_output)
        
        # --------------------- 单元格对比 --------------------- #
        self.progress_current_task.emit(f"开始对比sheet【{sheet1.title}】单元格")
        blank_row_flag = 0
        
        for row1 in range(2, max_row1 + 1):  # 从第2行开始对比（跳过标题行）
            # 进度更新
            progress_percent += step_progress
            if self.check_thread_running():
                return 0
                
            # 定期更新进度，避免UI卡顿
            if (row1 - 1) % self.update_frequency == 0:
                self.progress_updated.emit(round(progress_percent))
                self.progress_current_task.emit(
                    f"正在对比sheet【{sheet1.title}】第{row1}/{max_row1}行, 进度: {progress_percent:.1f}%"
                )
            
            # 获取目标行号
            row2 = index_column_mapping.get(row1, 0)
            if row2 == 0:
                blank_row_flag += 1
                row_changed_list[row1] = 3 #表征本行是否变更的flag，3：未匹配上的新增行，青色
                if blank_row_flag >= 20:
                    self.progress_current_task.emit("连续20行无匹配，结束对比")
                    break
                continue
                
            blank_row_flag = 0
            row_isChanged_status_flag = False
            
            # 逐列对比
            for col1 in range(1, max_col1 + 1):
                if self.check_thread_running():
                    return 0
                    
                sheet1_cell = sheet1.cell(row=row1, column=col1)
                sheet2_cell = sheet2.cell(row=row2, column=col1)
                
                if self.cell_consistency_check(sheet1_cell, sheet2_cell):
                    sheet1_cell.fill = PatternFill(
                        start_color=self.Agreed_color, 
                        end_color=self.Agreed_color, 
                        fill_type="solid"
                    )
                else:
                    row_isChanged_status_flag = True
                    sheet1_cell.fill = PatternFill(
                        start_color=self.Not_Agreed_color, 
                        end_color=self.Not_Agreed_color, 
                        fill_type="solid"
                    )
            if row_isChanged_status_flag:
                row_changed_list[row1] = 2  #表征本行是否变更的flag，2：对比不一致，红色
            else:
                row_changed_list[row1] = 1  #表征本行是否变更的flag，1：对比一致，绿色
        # 创建变更标记列
        self.create_row_changed_rows(sheet1, row_changed_list)
        end_time = time.time()
        end_time_output = f"单元格对比耗时: {end_time - row_mapping_time}s"
        print(end_time_output)
        
        self.result_info += textwrap.dedent(f"""
        参与对比行数:{len(row_changed_list)}
        变更行数:{sum(1 for v in row_changed_list.values() if v == 2)}
        新增行数:{sum(1 for v in row_changed_list.values() if v == 3)}
        """)
        
        self.progress_current_task.emit(textwrap.dedent(f"""
        ======================================
        {check_index_time_output}
        {split_time_output}
        {row_mapping_time_output}
        {end_time_output}
        ======================================
        """))
        self.progress_updated.emit(target_progress)
        return sheet1
    
    def mapping_col_by_title(self, sheet1, sheet2, title_row_number):
        """
        基于标题行在两个sheet之间建立列映射
        
        Args:
            sheet1: 第一个工作表对象
            sheet2: 第二个工作表对象
            title_row_number: 标题行的行号
            
        Returns:
            dict: 列映射字典，键为sheet1的列号，值为sheet2的列号
        """
        # 检查标题行号是否有效
        if not title_row_number or title_row_number > min(sheet1.max_row, sheet2.max_row):
            error = (
                f"mapping_col_by_title()方法中, 参数title_row_number = 【{title_row_number}】\n"
                f"1. 参数title_row_number不得为0\n"
                f"2. 参数title_row_number不得大于sheet行数【{min(sheet1.max_row, sheet2.max_row)}】\n"
            )
            raise ValueError(error)
        
        # 预索引第二个sheet的所有列标题
        # 键: 处理后的标题文本，值: 列号
        title_to_col_map: Dict[str, int] = {}
        
        max_col2 = sheet2.max_column
        for col2 in range(1, max_col2 + 1):
            if self.check_thread_running():
                return {}
                
            # 获取并处理标题文本
            title_text = str(sheet2.cell(row=title_row_number, column=col2).value)
            processed_text = self._process_title_text(title_text)
            
            if processed_text:
                title_to_col_map[processed_text] = col2
        
        # 建立列映射
        max_col1 = sheet1.max_column
        col_mapping: Dict[int, int] = {}  # 列映射结果
        blank_row_count = 0
        for col1 in range(1, max_col1 + 1):
            if self.check_thread_running():
                return {}
                
            # 更新进度
            if col1 % self.update_frequency == 0:
                self.progress_current_task.emit(
                    f"正在为sheet【{sheet1.title}】的第{col1}/{max_col1}列查找匹配标题"
                )
            
            # 获取并处理当前列的标题文本
            title_text = str(sheet1.cell(row=title_row_number, column=col1).value)
            processed_text = self._process_title_text(title_text)
            
            # 处理空标题
            if not processed_text:
                col_mapping[col1] = 0  # 标记为未匹配
                continue
            
            # 处理空行
            if not processed_text:
                blank_row_count += 1
                if blank_row_count >= 20:
                    col_mapping[col1] = 0
                    print_info = f"连续20列标题为空，结束sheet标题行的匹配"
                    print(print_info)
                    self.progress_current_task.emit(print_info)
                    break
                    
                # 标记为未匹配并设置颜色
                col_mapping[col1] = 0
                for row in range(1, sheet1.max_row + 1):
                    sheet1.cell(row=row, column=col1).fill = PatternFill(
                        start_color=self.No_match_color,
                        end_color=self.No_match_color,
                        fill_type="solid"
                    )
                continue
            else:
                blank_row_count = 0
            # 在预索引中查找匹配列
            if processed_text in title_to_col_map:
                col_mapping[col1] = title_to_col_map[processed_text]
            else:
                # 未找到匹配，标记为0并设置颜色
                col_mapping[col1] = 0
                for row in range(1, sheet1.max_row + 1):
                    sheet1.cell(row=row, column=col1).fill = PatternFill(
                        start_color=self.No_match_color,
                        end_color=self.No_match_color,
                        fill_type="solid"
                    )
        
        return col_mapping

    def _process_title_text(self, text: str) -> str:
        """处理标题文本，移除特殊字符和空白"""
        return text.replace('_x000D_', '').replace('\r', '').replace('\n', '').replace(' ', '')

    def mapping_row_by_index(self, sheet1, sheet2, sheet1_index_value_list, sheet2_index_value_list, title_row_number=0):
        """
        基于索引列建立行映射关系
        
        Args:
            sheet1: 第一个工作表
            sheet2: 第二个工作表
            sheet1_index_value_list: sheet1索引列号列表
            sheet2_index_value_list: sheet2索引列号列表
            title_row_number: 标题行号
            
        Returns:
            dict: 行映射字典（键：sheet1行号，值：sheet2行号）
        """
        max_row1, max_row2 = sheet1.max_row, sheet2.max_row
        max_col1, max_col2 = sheet1.max_column, sheet2.max_column

        # 过滤掉值为0的列索引
        sheet1_index_columns = [col for col in sheet1_index_value_list if col != 0]
        sheet2_index_columns = [col for col in sheet2_index_value_list if col != 0]
        
        if not sheet1_index_columns or not sheet2_index_columns:
            error = f"mapping_row_by_index()方法中, sheet1 or sheet2的索引为空"
            raise ValueError(error)  # 没有有效的索引列，返回空映射
        
        # 检查索引列有效性
        for sheet1_index in sheet1_index_value_list:
            if sheet1_index > max_col1:
                error = f"sheet1_index_value_list中存在超出sheet列数的值【{sheet1_index} > max_row1({max_col1})】"
                raise ValueError(error)
        for sheet2_index in sheet2_index_value_list:
            if sheet2_index > max_col2:
                error = f"sheet1_index_value_list中存在超出sheet列数的值【{sheet2_index} > max_row1({max_col2})】"
                raise ValueError(error)
        self.progress_current_task.emit(f"当前行数为：{inspect.currentframe().f_lineno}")
        
        # 预索引第二个sheet的所有行
        # 键: 合并后的索引值，值: 行号列表（处理重复值的情况）
        sheet2_index_map: Dict[str, List[int]] = defaultdict(list)
        
        max_row2 = sheet2.max_row
        for row2 in range(title_row_number+1, max_row2 + 1):
            # self.progress_current_task.emit(f"当前行数为：{inspect.currentframe().f_lineno}")
            if self.check_thread_running():
                return 0
            merged_text = self.get_merged_text(sheet2, row2, sheet2_index_columns)
            if merged_text:
                sheet2_index_map[merged_text].append(row2)
        
        # 建立行映射
        max_row1 = sheet1.max_row
        row_mapping: Dict[int, int] = {}  # 行映射结果
        blank_row_count = 0
        for row1 in range(title_row_number+1, max_row1 + 1):
            if self.check_thread_running():
                return 0
                
            # 更新进度
            if row1 % self.update_frequency == 0:
                self.progress_current_task.emit(
                    f"正在为sheet【{sheet1.title}】的第{row1}/{max_row1}行查找匹配项"
                )
            
            
            # 获取当前行的合并文本
            merged_text = self.get_merged_text(sheet1, row1, sheet1_index_columns)
            
            # 处理空行
            if not merged_text:
                blank_row_count += 1
                if blank_row_count >= 20:
                    row_mapping[row1] = 0
                    print_info = f"连续20行索引列值为空，结束sheet【{sheet1.title}】的匹配"
                    print(print_info)
                    self.progress_current_task.emit(print_info)
                    for row in range(row1-19, row1+1):
                        row_mapping.pop(row, None)
                        self.set_rows_color(sheet1, row1, self.None_color)
                    break
                    
                # 标记为未匹配并设置颜色
                row_mapping[row1] = 0
                self.set_rows_color(sheet1, row1, self.No_match_color)
                continue
            blank_row_count = 0

            # 在预索引中查找匹配行
            if merged_text in sheet2_index_map:
                # 取第一个匹配的行（处理重复值的情况）
                row_mapping[row1] = sheet2_index_map[merged_text][0]
            else:
                # 未找到匹配，标记为0并设置颜色
                row_mapping[row1] = 0
                self.set_rows_color(sheet1, row1, self.No_match_color)
        return row_mapping
    
    def set_rows_color(self, sheet, target_row, color):
        """
        设置整列的颜色
        
        Args:
            sheet1: 要设置的工作表对象
            target_row: 要设置的行
        Returns:
            None
        """
        try:
            # 检查目标行是否在有效范围内
            if target_row < 1 or target_row > sheet.max_row:
                raise ValueError(f"目标行({target_row})超出有效范围(1-{sheet.max_row})")
            
            # 检查目标行是否存在数据
            # row_data = [sheet.cell(row=target_row, column=col).value 
            #         for col in range(1, sheet.max_column + 1)]
            # if all(cell_value is None for cell_value in row_data):
            #     raise ValueError(f"目标行({target_row})数据为空")
            
            # 为整行填充颜色
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=target_row, column=col).fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"
                )
                
        except ValueError as e:
            # 捕获并重新抛出值错误，保留原始错误信息
            raise ValueError(f"填充行颜色失败: {str(e)}") from e
        except Exception as e:
            # 捕获其他类型的异常，提供更具体的错误信息
            raise ValueError(f"填充行颜色时发生未知错误: {str(e)}") from e
        
    def compare_excel_sheet_by_index_mapping_title(
        self, 
        sheet1, 
        sheet2, 
        title_name_list: List[str], 
        title_row_number: int, 
        file1_name: str,
        current_progress: int = 0, 
        target_progress: int = 100
    ) -> Optional[Union[bool, int]]:
        """
        基于索引和标题映射综合对比两个工作表
        
        Args:
            sheet1: 第一个工作表
            sheet2: 第二个工作表
            title_name_list: 标题名称列表
            title_row_number: 标题行号
            file1_name: 文件名
            current_progress: 当前进度百分比
            target_progress: 目标进度百分比
            
        Returns:
            sheet1: 对比后的工作表
            0: 任务被终止
        """
        self.result_info = "" #清空上次对比结果信息
        start_time = time.time()
        self.clear_all_conditional_formatting(sheet1)

        
        # 基础参数获取
        max_row1 = sheet1.max_row
        max_col1 = sheet1.max_column
        
        row_changed_list = {}
        """row_changed_list用于表示本行是否存在变更，
            0: 表示index未匹配上（无颜色填充），
            1: 表示行无变更（绿色填充），
            2: 表示行有变更（红色填充），
            3: 表示行新增（青色填充），
        """
        
        # 数据预处理
        title_name_list = list(dict.fromkeys(title_name_list))  # 去重并保留顺序
        delta_progress = target_progress - current_progress
        progress_percent = current_progress
        
        # --------------------- 标题行检查 --------------------- #
        self.progress_current_task.emit(f"正在检查标题行(第{title_row_number}行)重复值")
        try:
            title_row_values1 = [str(value) for value in next(sheet1.iter_rows(
                min_row=title_row_number, 
                max_row=title_row_number, 
                values_only=True
            ))]
            title_row_values2 = [str(value) for value in next(sheet2.iter_rows(
                min_row=title_row_number, 
                max_row=title_row_number, 
                values_only=True
            ))]
        except StopIteration:
            raise ValueError(f"标题行(第{title_row_number}行)数据为空")
        
        # 检查重复标题
        counter1 = Counter(title_row_values1)
        duplicates1 = {k: v for k, v in counter1.items() if v > 1 and k not in ["None", ""]}
        counter2 = Counter(title_row_values2)
        duplicates2 = {k: v for k, v in counter2.items() if v > 1 and k not in ["None", ""]}
        
        if duplicates1:
            raise ValueError(f"文件1标题行存在重复: {duplicates1}")
        if duplicates2:
            raise ValueError(f"文件2标题行存在重复: {duplicates2}")
        
        duplicates_time = time.time()
        duplicates_time_output = f"标题行检查耗时: {duplicates_time - start_time}s"
        print(duplicates_time_output)
        
        # --------------------- 索引列定位 --------------------- #
        self.progress_current_task.emit("正在定位索引列位置 (name -> value)")
        index_value_list_file1 = []
        index_value_list_file2 = []
        
        for title in title_name_list:
            try:
                index_value1 = title_row_values1.index(title)
                index_value2 = title_row_values2.index(title)
            except ValueError:
                raise ValueError(f"未找到标题 '{title}' 对应的列, title【{title}】的列不存在")
            
            index_value_list_file1.append(index_value1 + 1)
            index_value_list_file2.append(index_value2 + 1)
        
        print(f"文件1索引列: {index_value_list_file1}, 文件2索引列: {index_value_list_file2}")
        
        # --------------------- 索引列重复检查 --------------------- #
        self.progress_current_task.emit("正在检查索引列是否存在重复值")
        if self.check_index_repeat(sheet1, index_value_list_file1, title_row_number) == 0:
            return 0
        if self.check_index_repeat(sheet2, index_value_list_file2, title_row_number) == 0:
            return 0
        
        check_index_time = time.time()
        check_index_time_output = f"索引检查耗时: {check_index_time - duplicates_time}s"
        print(check_index_time_output)
        
        # --------------------- 合并单元格处理 --------------------- #
        self.progress_current_task.emit("正在拆分合并单元格")
        if not self.split_merged_cells(sheet1, index_value_list_file1):
            return 0
        if not self.split_merged_cells(sheet2, index_value_list_file2):
            return 0
        
        split_time = time.time()
        split_time_output = f"合并单元格拆分处理耗时: {split_time - check_index_time}s"
        print(split_time_output)
        
        # --------------------- 行映射 --------------------- #
        self.progress_current_task.emit("正在建立行映射关系")
        index_column_mapping = self.mapping_row_by_index(
            sheet1, 
            sheet2, 
            index_value_list_file1, 
            index_value_list_file2,
            title_row_number,
        )
        if index_column_mapping == 0:
            return 0
        
        step_progress = delta_progress / len(index_column_mapping) if len(index_column_mapping) > 0 else 0
        row_mapping_time = time.time()
        row_mapping_time_output = f"行映射建立耗时: {row_mapping_time - split_time}s"
        print(row_mapping_time_output)
        
        # --------------------- 列映射 --------------------- #
        self.progress_current_task.emit("正在建立列映射关系")
        title_row_mapping = self.mapping_col_by_title(sheet1, sheet2, title_row_number)
        if title_row_mapping == 0:
            return 0
        
        col_mapping_time = time.time()
        col_mapping_time_output = f"列映射耗时: {col_mapping_time - row_mapping_time}s"
        print(col_mapping_time_output)
        
        # --------------------- 单元格对比 --------------------- #
        self.progress_current_task.emit("开始单元格对比")
        blank_row_flag = 0
        
        for row1 in range(title_row_number + 1, len(index_column_mapping)+title_row_number+1):  # 从标题行下一行开始
            # 进度更新
            progress_percent += step_progress
            if self.check_thread_running():
                return 0
                
            if (row1 - title_row_number + 1) % self.update_frequency == 0:
                self.progress_updated.emit(round(progress_percent))
                self.progress_current_task.emit(
                    f"正在对比第{row1}/{max_row1}行, 进度: {progress_percent:.1f}%"
                )
            
            # 获取目标行号
            row2 = index_column_mapping.get(row1, 0)
            if row2 == 0:
                blank_row_flag += 1
                row_changed_list[row1] = 3 #表征本行是否变更的flag，3：未匹配上的新增行，青色
                continue
                
            blank_row_flag = 0
            row_isChanged_status_flag = False
            # 单元格对比
            for col1 in range(1, max_col1 + 1):
                col2 = title_row_mapping.get(col1, 0)
                if col2 == 0:
                    continue
                sheet1_cell = sheet1.cell(row=row1, column=col1)
                sheet2_cell = sheet2.cell(row=row2, column=col2)
                
                if self.cell_consistency_check(sheet1_cell, sheet2_cell):
                    sheet1_cell.fill = PatternFill(
                        start_color=self.Agreed_color, 
                        end_color=self.Agreed_color, 
                        fill_type="solid"
                    )
                else:
                    row_isChanged_status_flag = True
                    sheet1_cell.fill = PatternFill(
                        start_color=self.Not_Agreed_color, 
                        end_color=self.Not_Agreed_color, 
                        fill_type="solid"
                    )

            # 记录行变更状态
            if row_isChanged_status_flag:
                row_changed_list[row1] = 2  #表征本行是否变更的flag，2：对比不一致，红色
            else:
                row_changed_list[row1] = 1  #表征本行是否变更的flag，1：对比一致，绿色
        
        # 创建变更标记列
        self.create_row_changed_rows(sheet1, row_changed_list, title_row_number)
        end_time = time.time()
        end_time_output = f"单元格对比耗时: {end_time - col_mapping_time}s"
        print(end_time_output)
        
        self.result_info += textwrap.dedent(f"""
        参与对比行数:{len(row_changed_list)}
        变更行数:{sum(1 for v in row_changed_list.values() if v == 2)}
        新增行数:{sum(1 for v in row_changed_list.values() if v == 3)}
        """)
        
        self.progress_current_task.emit(textwrap.dedent(f"""
        ======================================
        {duplicates_time_output}
        {check_index_time_output}
        {split_time_output}
        {row_mapping_time_output}
        {col_mapping_time_output}
        {end_time_output}
        ======================================
        """))
        self.progress_updated.emit(target_progress)
        return sheet1

    def create_row_changed_rows(self, sheet, row_changed_list, title_row_number=0):
        """
        创建变更标记列，显示每行的变更状态
        
        Args:
            sheet: 工作表对象
            row_changed_list: 行变更状态字典
            title_row_number: 标题行号（用于设置列标题）
        """
        # 插入变更标记列
        sheet.insert_cols(1)

        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # 如果输入了表头行数，采用映射表头的方式对比，需要在表头行增加“变更点”
        if title_row_number:
            sheet.cell(row=title_row_number, column=1).value = "变更点"
            # 设置字体样式（黑色字体）
            sheet.cell(row=title_row_number, column=1).font = Font(
                name='Arial',
                size=12,
                bold=True,
                color="000000"  # 黑色字体
            )
            # 设置对齐方式
            sheet.cell(row=title_row_number, column=1).alignment = Alignment(
                horizontal='center',  # 水平居中
                vertical='center',    # 垂直居中
                wrap_text=True        # 自动换行
            )
            # 设置边框
            sheet.cell(row=title_row_number, column=1).border = thin_border
            
        
        # 填充变更标记
        for key, value in row_changed_list.items():
            if value == 1:  # 一致（绿色）
                sheet.cell(row=key, column=1).fill = PatternFill(
                    start_color=self.Agreed_color, 
                    end_color=self.Agreed_color, 
                    fill_type="solid"
                )
                sheet.cell(row=key, column=1).value = "一致"
            elif value == 2:  # 不一致（红色）
                sheet.cell(row=key, column=1).fill = PatternFill(
                    start_color=self.Not_Agreed_color, 
                    end_color=self.Not_Agreed_color, 
                    fill_type="solid"
                )
                sheet.cell(row=key, column=1).value = "差异点"
            elif value == 3:  # 新增（青色）
                sheet.cell(row=key, column=1).fill = PatternFill(
                    start_color=self.No_match_color, 
                    end_color=self.No_match_color, 
                    fill_type="solid"
                )
                sheet.cell(row=key, column=1).value = "新增"
            else:
                pass
            
            # 设置字体样式（黑色字体）
            sheet.cell(row=key, column=1).font = Font(
                name='Arial',
                size=12,
                bold=True,
                color="000000"  # 黑色字体
            )
            # 设置对齐方式
            sheet.cell(row=key, column=1).alignment = Alignment(
                horizontal='center',  # 水平居中
                vertical='center',    # 垂直居中
                wrap_text=True        # 自动换行
            )
            # 设置边框
            sheet.cell(row=key, column=1).border = thin_border

    def saving_file(self, wb1, output_path):
        """
        保存对比后的工作簿
        
        Args:
            wb1: 工作簿对象
            output_path: 输出路径
            
        Returns:
            int: 1表示成功，0表示失败
        """
        # 保存第一个工作簿，此时已包含对比和填充颜色后的结果
        try:
            print(f"saving file")
            self.progress_current_task.emit(f"对比完成，文件保存中···")
            wb1.save(output_path)
            print(f"file saved")
        except Exception as e:
            if isinstance(e, PermissionError):
                error = f"没有权限保存文件到指定路径，请检查文件权限设置。"
            elif isinstance(e, OSError) and "磁盘空间不足" in str(e):
                error = f"磁盘空间不足，无法保存文件，请清理磁盘空间后再试。"
            elif isinstance(e, FileNotFoundError):
                error = f"保存文件时文件路径不存在：{str(e)}"
                try:
                    os.mkdir(output_path.replace(".\\", ""))
                    error = f"文件夹 {output_path} 创建成功。"
                    wb1.save(output_path)
                except FileExistsError:
                    error = f"文件夹 {output_path} 已经存在。"
                except PermissionError:
                    error = f"没有权限创建文件夹 {output_path}。"
            else:
                error = f"保存文件时出现未知错误：{str(e)}"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            self.Progress_percent = 0
            self.progress_current_task.emit(f"对比完成，File1保存成功")
            return 0
            
        return 1

    def check_thread_running(self):
        """检查线程是否在运行（用于终止任务）"""
        if not self.is_running:
            # self.comparison_finished.emit()
            self.progress_current_task.emit(f"用户强制终止对比进程")
            return 1
        else:
            return 0