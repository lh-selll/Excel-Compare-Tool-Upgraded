
import inspect
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart,
    Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.title import Title
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side

class BorderStyle:
    def __init__(self):
        # 1. 粗边框（最粗）
        self.borderthick_border = Border(
            left=Side(style="thick"),  # thick表示粗边框
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )

        # 2. 中等粗边框
        self.bordermedium_border = Border(
            left=Side(style="medium"),  # medium表示中等粗细
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium")
        )

        # 3. 细边框（默认常用）
        self.borderthin_border = Border(
            left=Side(style="thin"),  # thin表示细边框
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 4. 顶部粗边框（混合样式）
        self.border_top_thick_border = Border(
            top=Side(style="thick"),
            left=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thin")
        )
        # 5. 左侧粗边框（混合样式）
        self.border_left_thick_border = Border(
            left=Side(style="thick"),
            top=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thin")
        )
        # 6. 右侧粗边框（混合样式）
        self.border_right_thick_border = Border(
            left=Side(style="thin"),
            top=Side(style="thin"),
            right=Side(style="thick"),
            bottom=Side(style="thin")
        )
        # 7. 底部粗边框（混合样式）
        self.border_bottom_thick_border = Border(
            left=Side(style="thin"),
            top=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thick")
        )

class ExcelChartManager:
    """Excel图表管理类，支持创建、配置和插入多种图表"""
    
    def __init__(self, workbook=None):
        """初始化图表管理器"""
        self.workbook = workbook if workbook else Workbook()
        self.current_sheet = self.workbook.active  # 默认使用当前活跃工作表
        # 定义不同粗细的边框样式
        self.border_style = BorderStyle()

    def set_sheet(self, sheet_name=None):
        """切换或创建目标工作表"""
        if sheet_name:
            if sheet_name in self.workbook.sheetnames:
                self.current_sheet = self.workbook[sheet_name]
            else:
                self.current_sheet = self.workbook.create_sheet(title=sheet_name)
        
        self.current_sheet.sheet_view.showGridLines = False
        return self  # 支持链式调用
    
    @staticmethod
    def get_report_by_first_column(sheet, flag):
        """
        获取第一列的报告数据
        :param sheet: 工作表对象
        :param flag: 标志位，
            0表示直接对比1，
            1表示按索引列对比，
            2表示按索引列和映射列对比
        :return: 报告数据列表
        """
        # 存储第一列数据
        first_column = []
        
        # 遍历第一列的所有行（min_col=1, max_col=1 表示只取第一列）
        for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
            # row是一个元组，包含当前行第一列的单元格值
            first_column.append(row[0])
            
        # 统计一致、差异点、新增、删除行的数量
        agreed_rows_number = sum(1 for item in first_column if item == "一致")
        not_agreed_rows_number = sum(1 for item in first_column if item == "差异点")
        unique_rows_number = sum(1 for item in first_column if item == "新增")
        delete_rows_number = sum(1 for item in first_column if item == "删除行")
        if flag == 0:
            data = [
                ["一致行数", agreed_rows_number     ],
                ["差异行数", not_agreed_rows_number ],
            ]
        else:
            data = [
                ["一致行数", agreed_rows_number     ],
                ["差异行数", not_agreed_rows_number ],
                ["新增行数", unique_rows_number     ],
                ["删除行数", delete_rows_number     ],
            ]

        return data

    def _set_pie_slice_colors(self, chart, colors):
        """通过直接操作XML根元素设置饼图颜色"""
        # 获取图表的XML根元素（兼容所有版本的核心方法）
        root = chart._element
        
        # 查找所有数据系列节点（<c:ser>）
        ns = {"c": CHART_NS}
        ser_nodes = root.findall(".//c:ser", namespaces=ns)
        
        if not ser_nodes:
            raise ValueError("未找到数据系列节点，无法设置颜色")
        
        # 处理第一个数据系列（饼图通常只有一个系列）
        ser_node = ser_nodes[0]
        
        # 查找所有数据点节点（<c:pt>）
        pt_nodes = ser_node.findall(".//c:pt", namespaces=ns)
        
        # 为每个数据点设置颜色
        for i, pt_node in enumerate(pt_nodes):
            if i < len(colors):
                # 创建颜色设置的XML结构
                sp_pr = Element(f"{{{CHART_NS}}}spPr")  # 图形属性
                solid_fill = Element(f"{{{CHART_NS}}}solidFill")  # 纯色填充
                srgb_clr = Element(f"{{{CHART_NS}}}srgbClr", val=colors[i])  # 颜色值
                
                # 组装XML结构
                solid_fill.append(srgb_clr)
                sp_pr.append(solid_fill)
                pt_node.append(sp_pr)

    def add_test_data(self, data: list):
        """添加测试数据（用于示例）"""
        
        for row in data:
            self.current_sheet.append(row)
        return self  # 支持链式调用

    def add_cell_value(self, row, column, value_list, border_style = None):

        """
        在指定行和列添加单元格内容
        :param row: 行号（从1开始）
        :param column: 列号（从1开始，1=A列，2=B列...）
        :param value: 要添加的值
        """
        print(f"row = {row}\ncolumn = {column}\nvalue_list = {value_list}")
        if border_style != None:
            if border_style == "thick":
                border_style = self.border_style.borderthick_border
            elif border_style == "medium":
                border_style = self.border_style.bordermedium_border
            elif border_style == "thin":
                border_style = self.border_style.borderthin_border
            elif border_style == "top":
                border_style = self.border_style.border_top_thick_border
            elif border_style == "left":
                border_style = self.border_style.border_left_thick_border
            elif border_style == "right":
                border_style = self.border_style.border_right_thick_border
            elif border_style == "bottom":
                border_style = self.border_style.border_bottom_thick_border
            else:
                border_style = None


        rows = row
        for row_data in value_list:
            col = column
            for data in row_data:
                self.current_sheet.cell(row=rows, column=col, value=data)
                self.current_sheet.cell(row=rows, column=col).border = border_style
                self.current_sheet.cell(row=rows, column=col).alignment = Alignment(wrap_text=True)    #把第一个文件的单元格设为自动换行
                col += 1
            rows += 1
            
    def set_sheet_main_title(self, row, col, height, width, content, title_type="main"):
        """
        设置工作表的主标题
        :param row: 行号（从1开始）
        :param col: 列号（从1开始，1=A列，2=B列...）
        :param height: 标题高度（合并的行数）
        :param width: 标题宽度（合并的列数）
        :param content: 标题内容
        :param title_type: 标题类型，"main"为主要标题，"sub"为副标题
        """
        #添加chart_sheet的主标题
        print(f"添加chart_sheet的主标题: row = {row}, col = {col}, height = {height}, width = {width}, content = {content}, title_type = {title_type}")

        self.current_sheet.merge_cells(start_row=row, start_column=col, end_row=row+height-1, end_column=col+width-1)
        self.add_cell_value(row, col, [[content]])

        print(f"content = {content}")
        
        if title_type == "main":
            # 1. 设置行高
            # 主标题行高（适当增大，突出标题）
            row_height = 60  # 行高值，数值越大行越高
            title_font = Font(
                name="微软雅黑",  # 选择清晰的无衬线字体
                size=12,         # 较大字号突出标题
                bold=True,       # 加粗
                color="1F4E78",  # 深蓝色，专业大气
                italic=False     # 不倾斜
            )
            border = self.border_style.borderthick_border
            fill = PatternFill(
                start_color="78c1e9",
                end_color="78c1e9",
                fill_type="solid"
            )
        elif title_type == "sub":
            row_height = 50  # 行高值，数值越大行越高
            title_font = Font(
                name="微软雅黑",  # 选择清晰的无衬线字体
                size=10,         # 较大字号突出标题
                bold=True,       # 加粗
                color="1F4E78",  # 深蓝色，专业大气
                italic=False     # 不倾斜
            )
            border = None
            fill = PatternFill(
                start_color="dbedf7",
                end_color="dbedf7",
                fill_type="solid"
            )

        else:
            row_height = 60  # 行高值，数值越大行越高
            title_font = Font(
                name="微软雅黑",  # 选择清晰的无衬线字体
                size=12,         # 较大字号突出标题
                bold=True,       # 加粗
                color="1F4E78",  # 深蓝色，专业大气
                italic=False     # 不倾斜
            )
            border = self.border_style.borderthick_border
            fill = PatternFill(
                start_color="78c1e9",
                end_color="78c1e9",
                fill_type="solid"
            )

        # 1. 设置行高
        # 主标题行高（适当增大，突出标题）
        self.current_sheet.row_dimensions[row].height = row_height  # 行高值，数值越大行越高
        
        # 2. 设置字体样式 - 大气美观的标题字体
        self.current_sheet.cell(row=row, column=col).font = title_font
        self.current_sheet.cell(row=row, column=col).fill = fill
        self.current_sheet.cell(row=row, column=col).font = title_font

        # 3. 设置单元格样式 - 左侧对齐
        self.current_sheet.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
        self.current_sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True)    #把第一个文件的单元格设为自动换行 

        # 4. 设置单元格样式 - 边框
        # 列标题用粗边框
        for i in range(row, row+height):
            for j in range(col, col+width):
                # 列标题用粗边框
                self.current_sheet.cell(row=i, column=j).border = border
                self.current_sheet.cell(row=i, column=j).alignment = Alignment(wrap_text=True)    #把第一个文件的单元格设为自动换行 
        
                
    def create_bar_chart(self, title, data_range, categories_range, pos="E2", show_labels=True, colors=None):
        """创建柱状图
        :param colors: 可选的颜色列表，如['FF0000', '00FF00', '0000FF']
        """
        print(f"data_range = {data_range}, categories_range = {categories_range}")
        chart = BarChart(orientation='vertical', grouping='standard')
        chart.title = title  # 直接使用字符串作为标题（兼容所有版本）
        chart.style = 10  # 预设样式
        # 调整图表大小以确保标签显示空间
        chart.width = 16  # 增加宽度
        chart.height = 8  # 增加高度
        # 调整柱子大小和间距，配合布局
        chart.barWidth = 15  # 柱子宽度
        chart.gapWidth = 60  # 类别间距
        
        # 添加数据系列
        chart.add_data(data_range, titles_from_data=True)
        
        # 设置分类轴（X轴）
        chart.set_categories(categories_range)
        # 确保坐标轴显示（默认已显示，这里显式设置）
        chart.x_axis.visible = True  # X轴可见
        chart.y_axis.visible = True  # Y轴可见
        
        # 设置坐标轴标题（直接用字符串）
        # 设置X轴（分类轴）标题，确保类别正确关联
        chart.x_axis.title = "类别"
        chart.x_axis.tickLblPos = "low"  # 标签显示在轴下方（关键设置）
        chart.x_axis.majorGridlines = None  # 移除网格线
        # chart.x_axis.tickMarkSkip = 1  # 强制显示所有标签
        # Y轴标题
        chart.y_axis.title = "数值"
        chart.y_axis.majorGridlines = None  # 移除网格线

        # 关键：设置图表布局和边距，控制内部坐标系大小
        # ManualLayout用于精确控制绘图区位置和大小
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.01,  # 左侧偏移（占图表宽度的20%），间接影响右侧留白
                y=0.01, # 顶部偏移（占图表高度的15%），间接影响底部留白
                layoutTarget="inner"  # 基于图表内部区域计算
            )
        )
        # 可选：显示数据标签（柱形上方显示数值）
        if show_labels:
            chart.dataLabels = DataLabelList()
            # 1. 显示数值（最常用）
            chart.dataLabels.showVal = True  # 显示数据值
            # 2. 显示系列名称（如"销量"、"增长率"）
            chart.dataLabels.showSerName = False  # 显示系列名称
            # 3. 显示类别名称（如"产品A"、"产品B"）
            chart.dataLabels.showCatName = False  # 不显示类别名称（默认）
            
        # 为每个类别的柱子设置颜色
        if chart.series:
            series = chart.series[0]
            # 为每个数据点创建DataPoint并应用颜色
            for i in range(len(colors)):
                try:
                    # 仅需指定索引即可创建DataPoint
                    dp = DataPoint(idx=i)
                    series.dPt.append(dp)
                    # 创建图形属性并设置填充颜色
                    gp = GraphicalProperties(solidFill=colors[i])
                    # 为数据点应用颜色
                    series.dPt[i].graphicalProperties = gp
                except:
                    break
        # 插入图表
        self.current_sheet.add_chart(chart, pos)
        return self

    def create_line_chart(self, title, data_range, categories_range, pos="E18", colors=None):
        """创建折线图
        :param colors: 可选的颜色列表，如['FF0000', '00FF00', '0000FF']
        """
        chart = LineChart()
        chart.title = title  # 直接使用字符串标题
        chart.style = 12
        chart.marker = True  # 显示数据点标记
        
        for idx, col in enumerate(range(data_range.min_col, data_range.max_col + 1)):
            series_data = Reference(
                self.current_sheet,
                min_col=col,
                min_row=data_range.min_row,
                max_row=data_range.max_row
            )
            series = Series(series_data, title_from_data=True)
            if colors and idx < len(colors):
                series.graphicalProperties.line.solidFill = colors[idx]
                series.graphicalProperties.line.width = 30000  # 设置线宽
            chart.append(series)
        
        chart.set_categories(categories_range)
        chart.x_axis.title = "类别"
        chart.y_axis.title = "数值"
        
        self.current_sheet.add_chart(chart, pos)
        return self
    
    @staticmethod
    def create_referencec_data(sheet, labels_col, data_col, min_row, max_row):

        """创建饼图数据范围"""
        # 定义饼图标签范围
        pie_labels_range = Reference(
            sheet,
            min_col=labels_col,
            min_row=min_row,
            max_row=max_row
        )
        # 定义饼图数据范围
        pie_data_range = Reference(
            sheet,
            min_col=data_col,
            min_row=min_row-1,
            max_row=max_row
        )
        return pie_labels_range, pie_data_range

    def create_pie_chart(self, title, data_range, labels_range, pos="E34"):
        """创建饼图
        :param colors: 可选的颜色列表，如['FF0000', '00FF00', '0000FF']
        """
        chart = PieChart()
        chart.title = title  # 直接使用字符串标题
        chart.style = 15
        print(f"当前行数为：{inspect.currentframe().f_lineno} create_pie_chart, data_range = {data_range}")
        
        # 添加数据系列
        series = Series(data_range, title_from_data=True)
        chart.series = [series]
        chart.set_categories(labels_range)

        # 显示数据标签（数值+百分比）
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        
        self.current_sheet.add_chart(chart, pos)
        return self

    def save(self, filename):
        """保存工作簿"""
        self.workbook.save(filename)
        print(f"文件已保存：{filename}")


def main():
    # 1. 初始化图表管理器
    chart_manager = ExcelChartManager()
    
    # 2. 创建并切换到目标工作表
    chart_manager.set_sheet("销售数据图表")
    
    # 3. 添加测试数据
    # 示例数据：销售额
    sales_data = [
        ["月份", "销售额", "利润", "成本"],
        ["一月", 100, 50, 30],
        ["二月", 120, 60, 40],
        ["三月", 150, 70, 50],
        ["四月", 130, 65, 45],
        ["五月", 160, 75, 55],
        ["六月", 180, 80, 60]
    ]
    chart_manager.add_test_data(sales_data)
    
    # 4. 定义数据范围（行列索引从1开始）
    data_range = Reference(
        chart_manager.current_sheet,
        min_col=2,  # B列
        min_row=2,  # 第2行
        max_col=4,  # D列
        max_row=7   # 第7行 (因为增加了数据行)
    )
    categories_range = Reference(
        chart_manager.current_sheet,
        min_col=1,  # A列
        min_row=2,
        max_row=7
    )
    pie_labels_range = Reference(
        chart_manager.current_sheet,
        min_col=1,
        min_row=2,
        max_row=4
    )
    pie_data_range = Reference(
        chart_manager.current_sheet,
        min_col=2,
        min_row=2,
        max_row=4
    )
    
    # 定义颜色列表 (ARGB格式，例如 'FFRRGGBB')
    bar_colors = ['FF0000', '00FF00', '0000FF']  # 红色、绿色、蓝色
    line_colors = ['FF0000', '00FF00', '0000FF'] # 红色、绿色、蓝色
    pie_colors = ['FFC000', 'FF0000', '00B050']  # 橙色、红色、绿色
    
    # 5. 创建图表
    chart_manager.create_bar_chart(
        title="月度销售数据柱状图",
        data_range=data_range,
        categories_range=categories_range,
        pos="E2",
        colors=bar_colors
    ).create_line_chart(
        title="月度销售数据折线图",
        data_range=data_range,
        categories_range=categories_range,
        pos="E18",
        colors=line_colors
    ).create_pie_chart(
        title="前3月销量占比饼图",
        data_range=pie_data_range,
        labels_range=pie_labels_range,
        pos="E34",
        colors=pie_colors
    )
    
    # 6. 保存文件
    chart_manager.save("销售数据图表示例.xlsx")

if __name__ == "__main__":
    main()